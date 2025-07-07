import threading
import tkinter as tk
import tkinter.ttk as tkk
from datetime import datetime, timedelta, time
from tkinter import filedialog, ttk
import time as tempo

import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side, Alignment
from openpyxl.utils import get_column_letter


# ==============================================================================
#       PASSO EXTRA: PEGAR PASTA QUE VAI SALVAR O EXCEL FINAL
# ==============================================================================
pasta_final_do_excel = ""
nome_arquivo = ""


def pasta_final():
    global pasta_final_do_excel
    pasta_final_do_excel = filedialog.askdirectory(
        title="Pasta para salvar o arquivo final",
    )
    if pasta_final_do_excel:
        var_pasta_final.set(pasta_final_do_excel)
    else:
        var_pasta_final.set("Nenhuma pasta selecionada")
    print(f"Caminho da pasta_final definido em {pasta_final_do_excel}")

    btn_formar_listas_de_contrato_e_meses.grid(row=5, column=0, pady=10, padx=10)


def salvar_arquivo_final():
    global nome_arquivo
    nome_arquivo = entry_nome_planilha.get().strip()
    if not nome_arquivo:
        nome_arquivo = "planilha_final.xlsx"
    nome_arquivo += ".xlsx"

    filtrar_dados()


# ==============================================================================
#       PASSO 1: PEGAR CAMINHO DA PASTA E LISTAS DE CONTRATOS E MESES
# ==============================================================================

# Variavel caminho da pasta
caminho = ""


# Função para o user escolher a pasta das planilhas
def selecionar_pasta():
    # Tornando a variavel caminho em global
    global caminho

    caminho = filedialog.askdirectory(
        title="Selecione uma pasta",
    )
    # Atualiza a variável com o caminho selecionado
    if caminho:  # Verifica se um caminho foi selecionado
        var_caminho.set(caminho)
    else:
        var_caminho.set("Nenhuma pasta selecionada")
    print(f"Caminho da pasta definido em {caminho}")


# ¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨
# Variaveis globais da função (abrir_pasta_e_pegar_listas_contratos_e_meses)
lista_contratos = []
lista_meses_organizada = []
caminho_planilha_resumo = ""
caminho_planilha_categorias = ""
planilhas_area = []
mes = ""
# ¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨¨


# Thread da função (abrir_pasta_e_pegar_listas_contratos_e_meses)
def thread_abrir_pasta_e_pegar_listas_contratos_e_meses():
    t1 = threading.Thread(target=abrir_pasta_e_pegar_listas_contratos_e_meses)
    t1.start()


# Funcao que pega os dados e os transfoma nas listas de contrato e de meses
def abrir_pasta_e_pegar_listas_contratos_e_meses():
    global lista_contratos, lista_meses_organizada, caminho_planilha_resumo, planilhas_area, caminho_planilha_categorias

    # Variavel lista meses de apoio
    lista_meses = []

    # Remoção e adição de interface gráfica
    btn_procurar_pasta.grid_forget()
    entry_caminho.grid_forget()
    entry_pasta_final.grid_forget()
    btn_formar_listas_de_contrato_e_meses.grid_forget()
    label_status.grid(row=0, column=0, pady=10, padx=10)
    pasta_final_label.grid_forget()
    btn_pasta_final.grid_forget()

    # Formando o caminho das planilhas
    for pasta, subpasta, arquivos in os.walk(caminho):
        for arquivo in arquivos:
            if arquivo.endswith(".xlsx"):
                planilha_para_planilha_areas = openpyxl.load_workbook(
                    os.path.join(pasta, arquivo), data_only=True
                )
                planilhas_area.append(planilha_para_planilha_areas)
                caminho_arquivo = os.path.join(pasta, arquivo)
                print(f"Arquivo: {caminho_arquivo}")

                # Abrindo planilha resumo
                if caminho_arquivo.endswith("RESUMO.xlsx"):
                    caminho_planilha_resumo = caminho_arquivo
                    planilha = openpyxl.load_workbook(caminho_arquivo)
                    aba = planilha.active

                    # Pegando a lista de contratos da planilha resumo
                    for coluna in aba.iter_cols(min_col=2, max_col=2, min_row=7):
                        for cel in coluna:
                            if cel.value:
                                lista_contratos.append(cel.value)
                    label_status.config(
                        text="Lista de contratos gerada com sucesso! Aguarde..."
                    )
                    tempo.sleep(1)
                elif caminho_arquivo.endswith("Classificação de equipe CB.xlsx"):
                    caminho_planilha_categorias = caminho_arquivo
                elif caminho_arquivo.endswith("Drenagem.xlsx"):
                    planilha = openpyxl.load_workbook(caminho_arquivo, data_only=True)
                    aba = planilha.active

                    # Pegando a lista de meses da planilha drenagem
                    for coluna in aba.iter_cols(min_col=5, min_row=5, max_row=5):
                        for cel in coluna:
                            if cel.value and isinstance(cel.value, str):
                                # Verificar se a string contém ' a '
                                if " a " in cel.value:
                                    data_inicio_str, data_final_str = cel.value.split(
                                        " a "
                                    )
                                    try:
                                        data_inicio = datetime.strptime(
                                            data_inicio_str, "%d/%m/%Y"
                                        )
                                        data_final = datetime.strptime(
                                            data_final_str, "%d/%m/%Y"
                                        )

                                        if data_final.day <= 2:
                                            # Considerar o mês anterior
                                            mes_ano = (
                                                data_final
                                                - timedelta(days=data_final.day)
                                            ).strftime("%m/%Y")
                                            lista_meses.append(mes_ano)
                                        else:
                                            # Considerar o mês da segunda data
                                            mes_ano = data_final.strftime("%m/%Y")
                                            lista_meses.append(mes_ano)
                                    except ValueError:
                                        print(f"Erro ao processar datas: {cel.value}")
                                else:
                                    print(f"Formato inválido na célula: {cel.value}")
                    label_status.config(
                        text="Lista de meses gerada com sucesso! Aguarde..."
                    )
                    tempo.sleep(1)
    label_status.config(text="Concluido!")
    tempo.sleep(1)
    thread_criar_janela_e_filtros()

    # Remover duplicatas e exibir listas finais
    # Organizar a lista de contratos em ordem alfabética
    lista_contratos = list(set(lista_contratos))
    lista_contratos.sort(key=str.lower)
    # print(f"Contratos: {lista_contratos}")

    # Organizar a lista de meses para ficar sem itens duplicados
    for item in lista_meses:
        if item not in lista_meses_organizada:
            lista_meses_organizada.append(item)
    lista_meses_organizada[:0] = ["Todos os meses"]
    # print(f"Meses: {lista_meses_organizada}")


# ==============================================================================
#              PASSO 2: CRIAR A JANELA DE FILTRO E AS COMBOBOXS
# ==============================================================================


# Thread da função (criar_janela_e_filtros)
def thread_criar_janela_e_filtros():
    t2 = threading.Thread(target=criar_janela_e_filtros)
    t2.start()


# Função para habilitar o botão de filtrar
def habilitar_botao(*args):
    if (
        caixa_lista_contratos.get() != "" and caixa_lista_meses.get() != ""
    ):  # Se ambas as comboboxes tiverem valor selecionado
        btn_filtrar.config(state=tk.NORMAL)  # Habilita o botão
    else:
        btn_filtrar.config(state=tk.DISABLED)  # Desabilita o botão


# Função para criar a janela de filtragem
def criar_janela_e_filtros():
    label_status.grid_forget()
    label_caixa_contratos.grid(row=0, column=0, padx=10, pady=10)
    caixa_lista_contratos.config(values=lista_contratos)
    caixa_lista_contratos.grid(row=1, column=0, pady=10, padx=10)
    label_caixa_meses.grid(row=0, column=1, padx=10, pady=10)
    caixa_lista_meses.config(values=lista_meses_organizada)
    caixa_lista_meses.grid(row=1, column=1, pady=10, padx=10)
    nome_planilha_label.grid(row=2, column=0, padx=10, pady=10)
    entry_nome_planilha.grid(row=2, column=1, padx=10, pady=10)
    btn_filtrar.grid(row=3, column=0, pady=10, padx=10)
    root.geometry("800x400+600+250")
    # frame.pack_forget()
    frame2.config(width=800, height=250)
    planilhas_area.pop()
    planilhas_area.pop()


# ==============================================================================
#          PASSO 3: FILTRAR OS DADOS DE ACORDO COM A ESCOLHA DO USER
# ==============================================================================


# Função para pegar a linha do contrato
def buscar_linha_contrato(caminho_resumo, contrato_selecionado):
    # Abrir a planilha RESUMO.xlsx
    planilha = openpyxl.load_workbook(caminho_resumo)
    aba = planilha.active

    # Percorrer a coluna de contratos (coluna 2, a partir da linha 7)
    for linha in range(7, aba.max_row + 1):
        contrato_cell = aba.cell(row=linha, column=2)  # Coluna 2 para contratos

        # Verificar se o valor da célula é o mesmo que o contrato selecionado
        if contrato_cell.value == contrato_selecionado:
            return linha  # Retorna a linha do contrato

    return None  # Retorna None se o contrato não for encontrado


# Função para filtrar os dados
def filtrar_dados():
    label_status_final.grid(row=4, column=0, padx=10, pady=10)
    global mes
    # Definindo variaveis para a linha do contrato e o mês escolhido
    linha_contrato = buscar_linha_contrato(
        caminho_planilha_resumo, caixa_lista_contratos.get()
    )
    mes = caixa_lista_meses.get()
    for planilha in planilhas_area:
        aba = planilha.active
        # print(aba)

    def numeros_de_funcionarios_por_semana_no_mes_escolhido(planilha):
        lista = []
        aba = planilha.active
        for coluna in aba.iter_cols(min_col=5, min_row=5, max_row=5):
            for celula in coluna:
                # print('celula', celula)
                if celula.value:
                    try:
                        data_inicio_str, data_final_str = celula.value.split(" a ")
                        data_final = datetime.strptime(
                            data_final_str.strip(), "%d/%m/%Y"
                        )

                        if data_final.day <= 2:
                            mes_ano_data = (
                                data_final - timedelta(days=data_final.day)
                            ).strftime("%m/%Y")
                        else:
                            mes_ano_data = data_final.strftime("%m/%Y")

                        if mes_ano_data == mes:
                            celulas_mescladas = 1
                            for celula_mesclada in aba.merged_cells.ranges:
                                #                                     print('Celula mesclada: ',celula_mesclada)
                                if celula.coordinate in celula_mesclada:
                                    # print('coordenadas: ', celula.coordinate)
                                    num_funcionarios_da_semana = (
                                        celula_mesclada.max_col
                                        - celula_mesclada.min_col
                                        + 1
                                    )
                                    # print(num_funcionarios_da_semana)
                                    lista.append(num_funcionarios_da_semana)
                    except:
                        print("erro")
        return lista

    def verificar_categoria(
        aba, coluna, num_func, i, data_final, linha_DO_CONTRATO, area
    ):
        #        print(f'Coluna = {coluna}')
        #        print(f'Linha = {linha_DO_CONTRATO}')
        dia_estagiario = "-"
        dia_junior = "-"
        dia_pleno = "-"
        dia_senior = "-"
        #        print(f'Data final = {data_final}')
        planilha_categoria = openpyxl.load_workbook(caminho_planilha_categorias)
        aba_categoria = planilha_categoria.active
        numero_de_funcionarios = num_func[i]
        #        print(f'Numero de func = {numero_de_funcionarios}')
        for qnt in range(numero_de_funcionarios):
            cargo = "-"
            #            print(f'Qnt = {qnt}')
            funcionario = aba.cell(row=6, column=coluna + qnt).value
            #            print(f'Funcionario = {funcionario}')
            for col in aba_categoria.iter_cols(min_col=2, max_col=2, min_row=4):
                for celula in col:
                    if celula.value == funcionario:
                        #                        print(f'Funcionario celula {celula.value}')
                        linha = celula.row

                        # Data do início (estagiário)
                        if aba_categoria.cell(row=linha, column=3).value:
                            dia_estagiario = aba_categoria.cell(
                                row=linha, column=3
                            ).value.date()
                        #                            print(f'Data estagiario = {dia_estagiario}')
                        else:
                            dia_estagiario = "-"
                        #                           print(f'Data estagiario = {dia_estagiario}')

                        # Data do início (junior)
                        if aba_categoria.cell(row=linha, column=4).value:
                            dia_junior = aba_categoria.cell(
                                row=linha, column=4
                            ).value.date()
                        #                            print(f'Data junior = {dia_junior}')
                        else:
                            dia_junior = "-"
                        #                            print(f'Data junior = {dia_junior}')

                        # Data do início (pleno)
                        if aba_categoria.cell(row=linha, column=5).value:
                            dia_pleno = aba_categoria.cell(
                                row=linha, column=5
                            ).value.date()
                        #                            print(f'TESTETESTETESTETESTETESTE {data_final >= dia_pleno}')
                        #                            print(f'Data pleno = {dia_pleno}')
                        else:
                            dia_pleno = "-"
                        #                            print(f'Data pleno = {dia_pleno}')

                        # Data início (senior)
                        if aba_categoria.cell(row=linha, column=6).value:
                            dia_senior = aba_categoria.cell(
                                row=linha, column=6
                            ).value.date()
                        #                            print(f'Data senior = {dia_senior}')
                        else:
                            dia_senior = "-"
            #                            print(f'Data senior = {dia_senior}')
            if dia_estagiario != "-":
                cargo = "estagiario"
            if dia_junior != "-":
                if data_final >= dia_junior:
                    cargo = "junior"
            if dia_pleno != "-":
                if data_final >= dia_pleno:
                    cargo = "pleno"
            if dia_senior != "-":
                if data_final >= dia_senior:
                    cargo = "senior"
            #            print(f'Cargo = {cargo}')
            #            print(f'VALOR DA CELULA DOS DIAS DOS FUNCIONARIOS!!!!!!!!!!! {aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).value}')
            #            print(f'VALOR DA COLUNA!!!!!!!!!!! {aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).column}')
            #            print(f'VALOR DA LINHA!!!!!!!!!!! {aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).row}')
            if cargo == "estagiario":
                if aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).value:
                    area[0] += aba.cell(
                        row=linha_DO_CONTRATO, column=coluna + qnt
                    ).value
            elif cargo == "junior":
                if aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).value:
                    area[1] += aba.cell(
                        row=linha_DO_CONTRATO, column=coluna + qnt
                    ).value
            elif cargo == "pleno":
                if aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).value:
                    area[2] += aba.cell(
                        row=linha_DO_CONTRATO, column=coluna + qnt
                    ).value
            else:
                if aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).value:
                    area[3] += aba.cell(
                        row=linha_DO_CONTRATO, column=coluna + qnt
                    ).value

        #            print(area[0])
        #            print(area[1])
        #            print(area[2])
        #            print(area[3])
        #            print(f'Area = {area}')
        return area

    def processar_planilha_mes_especifico():
        func_drenagem = numeros_de_funcionarios_por_semana_no_mes_escolhido(
            planilhas_area[0]
        )
        func_estrutura = numeros_de_funcionarios_por_semana_no_mes_escolhido(
            planilhas_area[1]
        )
        func_geologia = numeros_de_funcionarios_por_semana_no_mes_escolhido(
            planilhas_area[2]
        )
        func_geometria = numeros_de_funcionarios_por_semana_no_mes_escolhido(
            planilhas_area[3]
        )
        func_geotecnia = numeros_de_funcionarios_por_semana_no_mes_escolhido(
            planilhas_area[4]
        )
        func_pavimento = numeros_de_funcionarios_por_semana_no_mes_escolhido(
            planilhas_area[5]
        )
        func_sinalizacao = numeros_de_funcionarios_por_semana_no_mes_escolhido(
            planilhas_area[6]
        )
        func_terraplenagem = numeros_de_funcionarios_por_semana_no_mes_escolhido(
            planilhas_area[7]
        )
        print(f"Contrato escolhido: {caixa_lista_contratos.get()}")
        #        print(f'Funcionarios Drenagem: {func_drenagem} \n'
        #              f' Funcionarios estrutura: {func_estrutura} \n'
        #              f'Funcionarios geologia: {func_geologia} \n'
        #              f'Funcionarios geometria: {func_geometria} \n'
        #              f'Funcionarios geotecnia: {func_geotecnia} \n'
        #              f'Funcionarios pavimento: {func_pavimento} \n'
        #              f'Funcionarios sinalizacao: {func_sinalizacao} \n'
        #              f'Funcionarios terraplenagem: {func_terraplenagem}')

        # Definindo variaveis para os dias trabalhados por area e cargo
        drenagem = [0, 0, 0, 0]
        estrutura = [0, 0, 0, 0]
        geologia = [0, 0, 0, 0]
        geometria = [0, 0, 0, 0]
        geotecnia = [0, 0, 0, 0]
        pavimento = [0, 0, 0, 0]
        sinalizacao = [0, 0, 0, 0]
        terraplenagem = [0, 0, 0, 0]

        # Linha dos dias
        linha = buscar_linha_contrato(
            caminho_planilha_resumo, caixa_lista_contratos.get()
        )

        for planilha in planilhas_area:  # cada planilha das planilhas de area
            i = 0
            aba = planilha.active
            for col in aba.iter_cols(
                min_row=5, max_row=5, min_col=5
            ):  # colunas na linha das semanas
                for cel in col:
                    if cel.value:  # se tiver data na celula da linha das semanas
                        try:
                            data_inicio, data_final = cel.value.split(" a ")
                            data_final = datetime.strptime(
                                data_final.strip(), "%d/%m/%Y"
                            ).date()
                            # print(data_final)

                            if data_final.day <= 2:
                                mes_ano_data = (
                                    data_final - timedelta(days=data_final.day)
                                ).strftime("%m/%Y")
                                # print(mes_ano_data)
                            else:
                                mes_ano_data = data_final.strftime("%m/%Y")
                            # print(mes_ano_data)

                            if (
                                mes_ano_data == mes
                            ):  # ve se o mes achado corresponde com o mes pedido do user
                                coluna = (
                                    cel.column
                                )  # coluna que tem a celula com a data da semana
                                # funcao com argumentos (aba = aba da planilha de area atual, coluna = coluna com o dia da semana, e func_drenagem = lista com a quantidade de funcionarios por semana de drenagem)
                                if planilha is planilhas_area[0]:
                                    drenagem_list = verificar_categoria(
                                        aba,
                                        coluna,
                                        func_drenagem,
                                        i,
                                        data_final,
                                        linha,
                                        drenagem,
                                    )
                                    drenagem = drenagem_list
                                    # print(drenagem_list)
                                    i += 1
                                elif planilha is planilhas_area[1]:
                                    estrutura_list = verificar_categoria(
                                        aba,
                                        coluna,
                                        func_estrutura,
                                        i,
                                        data_final,
                                        linha,
                                        estrutura,
                                    )
                                    estrutura = estrutura_list
                                    i += 1
                                    # print(estrutura_list)
                                elif planilha is planilhas_area[2]:
                                    geologia_list = verificar_categoria(
                                        aba,
                                        coluna,
                                        func_geologia,
                                        i,
                                        data_final,
                                        linha,
                                        geologia,
                                    )
                                    geologia = geologia_list
                                    i += 1
                                    # print(geologia_list)
                                elif planilha is planilhas_area[3]:
                                    geometria_list = verificar_categoria(
                                        aba,
                                        coluna,
                                        func_geometria,
                                        i,
                                        data_final,
                                        linha,
                                        geometria,
                                    )
                                    geometria = geometria_list
                                    i += 1
                                    # print(geometria_list)
                                elif planilha is planilhas_area[4]:
                                    geotecnia_list = verificar_categoria(
                                        aba,
                                        coluna,
                                        func_geotecnia,
                                        i,
                                        data_final,
                                        linha,
                                        geotecnia,
                                    )
                                    geotecnia = geotecnia_list
                                    i += 1
                                    # print(geotecnia_list)
                                elif planilha is planilhas_area[5]:
                                    pavimento_list = verificar_categoria(
                                        aba,
                                        coluna,
                                        func_pavimento,
                                        i,
                                        data_final,
                                        linha,
                                        pavimento,
                                    )
                                    pavimento = pavimento_list
                                    i += 1
                                    # print(pavimento_list)
                                elif planilha is planilhas_area[6]:
                                    sinalizacao_list = verificar_categoria(
                                        aba,
                                        coluna,
                                        func_sinalizacao,
                                        i,
                                        data_final,
                                        linha,
                                        sinalizacao,
                                    )
                                    sinalizacao = sinalizacao_list
                                    i += 1
                                    # print(sinalizacao_list)
                                elif planilha is planilhas_area[7]:
                                    terraplenagem_list = verificar_categoria(
                                        aba,
                                        coluna,
                                        func_terraplenagem,
                                        i,
                                        data_final,
                                        linha,
                                        terraplenagem,
                                    )
                                    terraplenagem = terraplenagem_list
                                    i += 1
                                    # print(terraplenagem_list)
                        except:
                            continue
        print(f"LISTA FINAL DRENAGEM = {drenagem}")
        print(f"LISTA FINAL ESTRUCTURA = {estrutura}")
        print(f"LISTA FINAL GEOLOGIA = {geologia}")
        print(f"LISTA FINAL GEOMETRIA = {geometria}")
        print(f"LISTA FINAL GEOTECNIA = {geotecnia}")
        print(f"LISTA FINAL PAVIMENTO = {pavimento}")
        print(f"LISTA FINAL SINALIZACAO = {sinalizacao}")
        print(f"LISTA FINAL TERRAPLENAGEM = {terraplenagem}")

        planilha_final(
            drenagem,
            estrutura,
            geologia,
            geometria,
            geotecnia,
            pavimento,
            sinalizacao,
            terraplenagem,
        )

    def verificar_categoria_todos(aba, coluna, i, data_final, linha_DO_CONTRATO, area):
        # print(f'Coluna = {coluna}')
        # print(f'Linha = {linha_DO_CONTRATO}')
        dia_estagiario = "-"
        dia_junior = "-"
        dia_pleno = "-"
        dia_senior = "-"
        # print(f'Data final = {data_final}')
        planilha_categoria = openpyxl.load_workbook(caminho_planilha_categorias)
        aba_categoria = planilha_categoria.active

        celula_mesclada = 1
        for celula_mesclada in aba.merged_cells.ranges:
            if aba.cell(row=5, column=coluna).coordinate in celula_mesclada:
                numero_de_funcionarios = (
                    celula_mesclada.max_col - celula_mesclada.min_col + 1
                )

        # print(f'Numero de func = {numero_de_funcionarios}')
        for qnt in range(numero_de_funcionarios):
            cargo = "-"
            # print(f'Qnt = {qnt}')
            funcionario = aba.cell(row=6, column=coluna + qnt).value
            # print(f'Funcionario = {funcionario}')
            for col in aba_categoria.iter_cols(min_col=2, max_col=2, min_row=4):
                for celula in col:
                    if celula.value == funcionario:
                        # print(f'Funcionario celula {celula.value}')
                        linha = celula.row

                        # Data do início (estagiário)
                        if aba_categoria.cell(row=linha, column=3).value:
                            dia_estagiario = aba_categoria.cell(
                                row=linha, column=3
                            ).value.date()
                            # print(f'Data estagiario = {dia_estagiario}')
                        else:
                            dia_estagiario = "-"
                            # print(f'Data estagiario = {dia_estagiario}')

                        # Data do início (junior)
                        if aba_categoria.cell(row=linha, column=4).value:
                            dia_junior = aba_categoria.cell(
                                row=linha, column=4
                            ).value.date()
                            # print(f'Data junior = {dia_junior}')
                        else:
                            dia_junior = "-"
                            # print(f'Data junior = {dia_junior}')

                        # Data do início (pleno)
                        if aba_categoria.cell(row=linha, column=5).value:
                            dia_pleno = aba_categoria.cell(
                                row=linha, column=5
                            ).value.date()
                            # print(f'TESTETESTETESTETESTETESTE {data_final >= dia_pleno}')
                            # print(f'Data pleno = {dia_pleno}')
                        else:
                            dia_pleno = "-"
                            # print(f'Data pleno = {dia_pleno}')

                        # Data início (senior)
                        if aba_categoria.cell(row=linha, column=6).value:
                            dia_senior = aba_categoria.cell(
                                row=linha, column=6
                            ).value.date()
                            # print(f'Data senior = {dia_senior}')
                        else:
                            dia_senior = "-"
                            # print(f'Data senior = {dia_senior}')
            if dia_estagiario != "-":
                cargo = "estagiario"
            if dia_junior != "-":
                if data_final >= dia_junior:
                    cargo = "junior"
            if dia_pleno != "-":
                if data_final >= dia_pleno:
                    cargo = "pleno"
            if dia_senior != "-":
                if data_final >= dia_senior:
                    cargo = "senior"
            # print(f'Cargo = {cargo}')
            # print(f'VALOR DA CELULA DOS DIAS DOS FUNCIONARIOS!!!!!!!!!!! {aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).value}')
            # print(f'VALOR DA COLUNA!!!!!!!!!!! {aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).column}')
            # print(f'VALOR DA LINHA!!!!!!!!!!! {aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).row}')
            if cargo == "estagiario":
                if aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).value:
                    area[0] += aba.cell(
                        row=linha_DO_CONTRATO, column=coluna + qnt
                    ).value
            elif cargo == "junior":
                if aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).value:
                    area[1] += aba.cell(
                        row=linha_DO_CONTRATO, column=coluna + qnt
                    ).value
            elif cargo == "pleno":
                if aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).value:
                    area[2] += aba.cell(
                        row=linha_DO_CONTRATO, column=coluna + qnt
                    ).value
            else:
                if aba.cell(row=linha_DO_CONTRATO, column=coluna + qnt).value:
                    area[3] += aba.cell(
                        row=linha_DO_CONTRATO, column=coluna + qnt
                    ).value

        #            print(area[0])
        #            print(area[1])
        #            print(area[2])
        #            print(area[3])
        #            print(f'Area = {area}')
        return area

    def processar_todos_os_meses():
        print(f"Contrato escolhido: {caixa_lista_contratos.get()}")

        # Definindo variaveis para os dias trabalhados por area e cargo
        drenagem = [0, 0, 0, 0]
        estrutura = [0, 0, 0, 0]
        geologia = [0, 0, 0, 0]
        geometria = [0, 0, 0, 0]
        geotecnia = [0, 0, 0, 0]
        pavimento = [0, 0, 0, 0]
        sinalizacao = [0, 0, 0, 0]
        terraplenagem = [0, 0, 0, 0]

        # Linha dos dias
        linha = buscar_linha_contrato(
            caminho_planilha_resumo, caixa_lista_contratos.get()
        )

        for planilha in planilhas_area:  # cada planilha das planilhas de area
            i = 0
            aba = planilha.active
            for col in aba.iter_cols(
                min_row=5, max_row=5, min_col=5
            ):  # colunas na linha das semanas
                for cel in col:
                    if cel.value:  # se tiver data na celula da linha das semanas
                        try:
                            data_inicio, data_final = cel.value.split(" a ")
                            data_final = datetime.strptime(
                                data_final.strip(), "%d/%m/%Y"
                            ).date()
                            # print(data_final)

                            coluna = (
                                cel.column
                            )  # coluna que tem a celula com a data da semana
                            # funcao com argumentos (aba = aba da planilha de area atual, coluna = coluna com o dia da semana, e func_drenagem = lista com a quantidade de funcionarios por semana de drenagem)
                            if planilha is planilhas_area[0]:
                                drenagem_list = verificar_categoria_todos(
                                    aba, coluna, i, data_final, linha, drenagem
                                )
                                drenagem = drenagem_list
                                # print(drenagem_list)
                                i += 1
                            elif planilha is planilhas_area[1]:
                                estrutura_list = verificar_categoria_todos(
                                    aba, coluna, i, data_final, linha, estrutura
                                )
                                estrutura = estrutura_list
                                i += 1
                                # print(estrutura_list)
                            elif planilha is planilhas_area[2]:
                                geologia_list = verificar_categoria_todos(
                                    aba, coluna, i, data_final, linha, geologia
                                )
                                geologia = geologia_list
                                i += 1
                                # print(geologia_list)
                            elif planilha is planilhas_area[3]:
                                geometria_list = verificar_categoria_todos(
                                    aba, coluna, i, data_final, linha, geometria
                                )
                                geometria = geometria_list
                                i += 1
                                # print(geometria_list)
                            elif planilha is planilhas_area[4]:
                                geotecnia_list = verificar_categoria_todos(
                                    aba, coluna, i, data_final, linha, geotecnia
                                )
                                geotecnia = geotecnia_list
                                i += 1
                                # print(geotecnia_list)
                            elif planilha is planilhas_area[5]:
                                pavimento_list = verificar_categoria_todos(
                                    aba, coluna, i, data_final, linha, pavimento
                                )
                                pavimento = pavimento_list
                                i += 1
                                # print(pavimento_list)
                            elif planilha is planilhas_area[6]:
                                sinalizacao_list = verificar_categoria_todos(
                                    aba, coluna, i, data_final, linha, sinalizacao
                                )
                                sinalizacao = sinalizacao_list
                                i += 1
                                # print(sinalizacao_list)
                            elif planilha is planilhas_area[7]:
                                terraplenagem_list = verificar_categoria_todos(
                                    aba, coluna, i, data_final, linha, terraplenagem
                                )
                                terraplenagem = terraplenagem_list
                                i += 1
                                # print(terraplenagem_list)
                        except:
                            continue
        print(f"LISTA FINAL DRENAGEM = {drenagem}")
        print(f"LISTA FINAL ESTRUCTURA = {estrutura}")
        print(f"LISTA FINAL GEOLOGIA = {geologia}")
        print(f"LISTA FINAL GEOMETRIA = {geometria}")
        print(f"LISTA FINAL GEOTECNIA = {geotecnia}")
        print(f"LISTA FINAL PAVIMENTO = {pavimento}")
        print(f"LISTA FINAL SINALIZACAO = {sinalizacao}")
        print(f"LISTA FINAL TERRAPLENAGEM = {terraplenagem}")

        planilha_final(
            drenagem,
            estrutura,
            geologia,
            geometria,
            geotecnia,
            pavimento,
            sinalizacao,
            terraplenagem,
        )

    # Se o mês escolhido for 'Todos os meses'
    if mes == "Todos os meses":
        # Se a opção for 'Todos os meses', chamamos a função de todos os meses
        thread_todos_os_meses = threading.Thread(target=processar_todos_os_meses)
        thread_todos_os_meses.start()
    # Se for um mês específico
    else:
        # Caso contrário, chamamos a função para o mês específico
        thread_mes_especifico = threading.Thread(
            target=processar_planilha_mes_especifico
        )
        thread_mes_especifico.start()


# ========================================================================================
#          PASSO 4: FUNÇÃO PARA GERAR UMA PLANILHA EM EXCEL COM OS DADOS OBTIDOS
# ========================================================================================


def planilha_final(
    drenagem,
    estrutura,
    geologia,
    geometria,
    geotecnia,
    pavimento,
    sinalizacao,
    terraplenagem,
):
    # Criação do arquivo e aba
    planilha_final = openpyxl.Workbook()
    aba_final = planilha_final.active
    aba_final.title = "Resumo"

    # Configuração das áreas (mantendo a sua estrutura original)
    areas_config = [
        {"nome": "Drenagem", "coluna_inicial": 3, "dados": drenagem},
        {"nome": "Estruturas", "coluna_inicial": 7, "dados": estrutura},
        {"nome": "Geologia", "coluna_inicial": 11, "dados": geologia},
        {"nome": "Geometria", "coluna_inicial": 15, "dados": geometria},
        {"nome": "Geotecnia", "coluna_inicial": 19, "dados": geotecnia},
        {"nome": "Pavimento", "coluna_inicial": 23, "dados": pavimento},
        {"nome": "Sinalização", "coluna_inicial": 27, "dados": sinalizacao},
        {"nome": "Terraplenagem", "coluna_inicial": 31, "dados": terraplenagem},
    ]

    # =========================================================================
    # 1. CONFIGURAÇÃO DE ESTILOS
    # =========================================================================
    # Alinhamento centralizado
    alinhamento_central = Alignment(horizontal="center", vertical="center")

    # Cores
    cinza = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    cinza_claro = PatternFill(
        start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
    )
    azul = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    azul_claro = PatternFill(
        start_color="B8CCE4", end_color="B8CCE4", fill_type="solid"
    )

    # Borda padrão
    borda_fina = Side(border_style="thin", color="000000")
    borda_completa = Border(
        left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina
    )

    # =========================================================================
    # 2. CABEÇALHO PRINCIPAL
    # =========================================================================
    # Células de título
    aba_final["A1"] = "Mês"
    aba_final["B1"] = "Contrato"

    # Aplicar borda e formatação em A3 e B3 ANTES de mesclar
    for col in [1, 2]:  # Colunas A (1) e B (2)
        celula = aba_final.cell(row=3, column=col)
        celula.value = "Dias" if col == 1 else ""  # Valor só na primeira célula
        celula.fill = cinza
        celula.alignment = alinhamento_central
        celula.border = borda_completa

    # Mesclar após aplicar bordas
    aba_final.merge_cells("A3:B3")

    # Formatação
    for celula in ["A1", "B1"]:
        aba_final[celula].fill = azul
        aba_final[celula].alignment = alinhamento_central
        aba_final[celula].border = borda_completa

    # Células de valores
    aba_final["A2"] = mes
    aba_final["B2"] = caixa_lista_contratos.get()
    for celula in ["A2", "B2"]:
        aba_final[celula].fill = azul_claro
        aba_final[celula].alignment = alinhamento_central
        aba_final[celula].border = borda_completa

    # =========================================================================
    # 3. FORMATAÇÃO DAS ÁREAS
    # =========================================================================
    categorias = ["Estagiario", "Junior", "Pleno", "Senior"]
    for area in areas_config:
        col = area["coluna_inicial"]

        # Título da área (linha 1)
        aba_final.merge_cells(
            start_row=1, start_column=col, end_row=1, end_column=col + 3
        )
        celula_titulo = aba_final.cell(row=1, column=col)
        celula_titulo.value = area["nome"]
        celula_titulo.fill = cinza
        celula_titulo.alignment = alinhamento_central
        celula_titulo.border = borda_completa

        # Cabeçalhos de cargos (linha 2)
        for i in range(4):
            celula = aba_final.cell(row=2, column=col + i)
            celula.value = categorias[i]
            celula.fill = cinza_claro
            celula.alignment = alinhamento_central
            celula.border = borda_completa

    # =========================================================================
    # 4. PREENCHIMENTO DE DADOS
    # =========================================================================
    linha = 3
    for area in areas_config:
        col = area["coluna_inicial"]
        for idx, valor in enumerate(area["dados"]):
            celula = aba_final.cell(row=linha, column=col + idx)
            celula.value = valor
            celula.alignment = alinhamento_central
            celula.border = borda_completa

    # =========================================================================
    # 4.1. LINHA DE HORAS (NOVO)
    # =========================================================================

    linha_horas = 4
    for area in areas_config:
        col = area["coluna_inicial"]
        for idx in range(4):
            celula = aba_final.cell(row=linha_horas, column=col + idx)
            celula.value = f"={get_column_letter(col + idx)}{linha}*8"  # Fórmula
            celula.number_format = "0"  # Formato numérico
            celula.alignment = alinhamento_central
            celula.border = borda_completa
    for col in [1, 2]:
        celula = aba_final.cell(row=4, column=col)
        celula.value = "Horas" if col == 1 else ""
        celula.fill = cinza
        celula.alignment = alinhamento_central
        celula.border = borda_completa
    aba_final.merge_cells("A4:B4")

    # =========================================================================
    # 5. LINHAS DE VALOR E CUSTO (MODIFICADO)
    # =========================================================================

    # Custo por Dia (linha 5)
    for col in [1, 2]:
        celula = aba_final.cell(row=5, column=col)
        celula.value = "Custo por hora" if col == 1 else ""
        celula.fill = cinza
        celula.alignment = alinhamento_central
        celula.border = borda_completa
    aba_final.merge_cells("A5:B5")

    # Preencher custos (primeira área vazia, outras com fórmula)
    for area_idx, area in enumerate(areas_config):
        col = area["coluna_inicial"]
        for i in range(4):
            celula = aba_final.cell(row=5, column=col + i)
            if area_idx == 0:  # Primeira área
                celula.number_format = '"R$ "#,##0.00'
                celula.value = None  # Célula em branco
            else:  # Demais áreas
                ref_col = areas_config[0]["coluna_inicial"] + i
                celula.value = f"={get_column_letter(ref_col)}5"
                celula.number_format = '"R$ "#,##0.00'
            celula.fill = cinza_claro
            celula.border = borda_completa

    # Total Custo (linha 6)
    for col in [1, 2]:
        celula = aba_final.cell(row=6, column=col)
        celula.value = "Total Custo" if col == 1 else ""
        celula.fill = cinza
        celula.alignment = alinhamento_central
        celula.border = borda_completa
    aba_final.merge_cells("A6:B6")

    for area in areas_config:
        col = area["coluna_inicial"]
        for i in range(4):
            celula = aba_final.cell(row=6, column=col + i)
            celula.value = (
                f"={get_column_letter(col + i)}4*{get_column_letter(col + i)}5"
            )
            celula.number_format = '"R$ "#,##0.00'
            celula.fill = cinza_claro
            celula.border = borda_completa

    # Linha 7 vazia sem bordas
    for col in range(1, aba_final.max_column + 1):
        aba_final.cell(row=7, column=col).border = Border()

    # =========================================================================
    # 6. TABELA DE TOTAIS (MODIFICADO)
    # =========================================================================
    # Célula Total (linha 8)
    for col in range(3, 7):
        celula = aba_final.cell(row=8, column=col)
        celula.value = "Total" if col == 3 else ""
        celula.fill = cinza
        celula.alignment = alinhamento_central
        celula.border = borda_completa
    aba_final.merge_cells("C8:F8")

    # Categorias (linha 9)
    for idx, categoria in enumerate(categorias):
        celula = aba_final.cell(row=9, column=3 + idx)
        celula.value = categoria
        celula.fill = cinza_claro
        celula.alignment = alinhamento_central
        celula.border = borda_completa

    # Totais (linha 10) com fórmulas
    for i in range(4):
        formula = f'=SUM({",".join([f"{get_column_letter(area["coluna_inicial"] + i)}6" for area in areas_config])})'
        celula = aba_final.cell(row=10, column=3 + i)
        celula.value = formula
        celula.number_format = '"R$ "#,##0.00'
        celula.alignment = alinhamento_central
        celula.border = borda_completa

    # =========================================================================
    # 7. AJUSTES FINAIS (MODIFICADO)
    # =========================================================================

    # Largura das colunas
    aba_final.column_dimensions["B"].width = 40
    aba_final.column_dimensions["A"].width = 20
    for col in range(1, aba_final.max_column + 1):
        if col != 2 and col != 1:  # Mantém a coluna B com 40, outras com 12
            aba_final.column_dimensions[get_column_letter(col)].width = 12

    # Altura das linhas
    for row in range(1, 10):
        aba_final.row_dimensions[row].height = 30

    aba_final.sheet_view.showGridLines = False

    planilha_final.save(f"{pasta_final_do_excel}/{nome_arquivo}")
    label_status_final.config(text="Arquivo gerado com sucesso!")
    tempo.sleep(1)
    label_status_final.grid_forget()


# ===================================
# = TK INTER
# ===================================

root = tk.Tk()
root.title("Selecionar Pasta")
root.geometry("700x400+600+250")

# Primeira, pedir pasta
frame = tk.Frame(root, width=500, height=50)
frame.pack()

frame2 = tk.Frame(root, width=500, height=150)
frame2.pack()

var_caminho = tk.StringVar()
var_pasta_final = tk.StringVar()

planilhas_label = tk.Label(frame2, text="Selecione a pasta com as planilhas")
planilhas_label.grid(row=0, column=0, padx=10, pady=10)

btn_procurar_pasta = tk.Button(
    frame2, command=selecionar_pasta, text="Selecionar pasta", bg="#4CAF50", fg="white"
)
btn_procurar_pasta.grid(row=1, column=0, pady=10, padx=10)

pasta_final_label = tk.Label(
    frame2, text="Selecione a pasta para salvar a planilha final"
)
pasta_final_label.grid(row=2, column=0, padx=10, pady=10)

btn_pasta_final = tk.Button(
    frame2, text="Selecionar pasta", command=pasta_final, bg="#4CAF50", fg="white"
)
btn_pasta_final.grid(row=3, column=0, padx=10, pady=10)

entry_caminho = tk.Entry(
    frame2,
    textvariable=var_caminho,
    width=45,
    state="readonly",
    readonlybackground="white",
)
entry_caminho.grid(row=1, column=1, pady=10, padx=10)

entry_pasta_final = tk.Entry(
    frame2,
    textvariable=var_pasta_final,
    width=45,
    state="readonly",
    readonlybackground="white",
)
entry_pasta_final.grid(row=3, column=1, pady=10, padx=10)


# Fim janela inicial
# --------------------------------------------------------------------------------------------------------------------------
# Segunda janela, captação de dados (lista_de_contratos e lista_de_meses)
btn_formar_listas_de_contrato_e_meses = tk.Button(
    frame2,
    command=thread_abrir_pasta_e_pegar_listas_contratos_e_meses,
    text="Processar",
    bg="#4CAF50",
    fg="white",
)

label_status = tkk.Label(frame2, text="Processando...")

barra_progresso = tk.ttk.Progressbar(
    frame2, orient="horizontal", length=300, mode="determinate", maximum=100
)
# Fim da segunda janela
# --------------------------------------------------------------------------------------------------------------------------
# Terceira janela, criação da janela filtros

label_caixa_contratos = tk.Label(frame2, text="Contrato: ")
caixa_lista_contratos = ttk.Combobox(
    frame2, values=lista_contratos, state="readonly", width=45
)

label_caixa_meses = tk.Label(frame2, text="Mês: ")
caixa_lista_meses = ttk.Combobox(
    frame2, values=lista_meses_organizada, state="readonly", width=15
)

caixa_lista_contratos.bind("<<ComboboxSelected>>", habilitar_botao)
caixa_lista_meses.bind("<<ComboboxSelected>>", habilitar_botao)

nome_planilha_label = tk.Label(
    frame2, text="Digite o nome que deseja para a planilha final: "
)

entry_nome_planilha = tk.Entry(frame2, width=45)

btn_filtrar = tk.Button(
    frame2, text="Filtrar", state=tk.DISABLED, command=salvar_arquivo_final
)

label_status_final = tkk.Label(frame2, text="Processando...")


root.mainloop()
