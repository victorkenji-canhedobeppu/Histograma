# -*- coding: utf-8 -*-

import pandas as pd
import tkinter as tk
import tkinter.font as tkFont
from tkinter import ttk, messagebox, simpledialog, filedialog
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.colors import SchemeColor
from openpyxl.drawing.fill import SolidColorFillProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import (
    GradientFillProperties,
    GradientStop,
    LinearShadeProperties,
)
from openpyxl.drawing.geometry import Bevel, Shape3D

import datetime
import openpyxl
from functools import partial
from datetime import datetime as datetimeFootHolder
import traceback
import json
import os

from config.settings import CARGOS

ESTADO_APP_FILE = "estado_app.json"


# MODIFICADO: A janela de diálogo agora inclui a seleção de disciplina.
class CustomDialog(tk.Toplevel):
    def __init__(self, parent, disciplinas):
        super().__init__(parent)
        self.transient(parent)
        self.title("Novo Funcionário")
        self.parent = parent
        self.result = None
        self.configure(bg="#F0F2F5")
        body = ttk.Frame(self, padding="10")
        self.initial_focus = self.body(body, disciplinas)
        body.pack(padx=10, pady=10)
        self.buttonbox()
        self.grab_set()
        if not self.initial_focus:
            self.initial_focus = self
        self.protocol("WM_DELETE_WINDOW", self.cancel)
        self.geometry(f"+{parent.winfo_rootx()+50}+{parent.winfo_rooty()+50}")
        self.initial_focus.focus_set()
        self.wait_window(self)

    def body(self, master, disciplinas):
        ttk.Label(master, text="Nome:").grid(
            row=0, column=0, sticky=tk.W, padx=5, pady=5
        )
        self.nome_entry = ttk.Entry(master, width=30)
        self.nome_entry.grid(row=0, column=1, padx=5, pady=5)

        # O CAMPO DE CARGO FOI REMOVIDO DAQUI

        ttk.Label(master, text="Disciplina Principal:").grid(
            row=1, column=0, sticky=tk.W  # A linha agora é 1
        )
        self.disciplina_combo = ttk.Combobox(
            master, values=disciplinas, state="readonly", width=28
        )
        self.disciplina_combo.grid(row=1, column=1, padx=5, pady=5)  # A linha agora é 1
        if disciplinas:
            self.disciplina_combo.current(0)

        return self.nome_entry

    def buttonbox(self):
        box = ttk.Frame(self)
        w = ttk.Button(
            box, text="OK", width=10, command=self.ok, style="Accent.TButton"
        )
        w.pack(side=tk.LEFT, padx=5, pady=5)
        w = ttk.Button(box, text="Cancelar", width=10, command=self.cancel)
        w.pack(side=tk.LEFT, padx=5, pady=5)
        self.bind("<Return>", self.ok)
        self.bind("<Escape>", self.cancel)
        box.pack(pady=5)

    def ok(self, event=None):
        nome = self.nome_entry.get().strip()
        disciplina = self.disciplina_combo.get()

        if not nome or not disciplina:
            messagebox.showwarning(
                "Entrada Inválida",
                "Nome e disciplina são obrigatórios.",
                parent=self,
            )
            return

        self.result = (nome, disciplina)  # Retorna a tupla com 2 elementos
        self.withdraw()
        self.update_idletasks()
        self.cancel()

    def cancel(self, event=None):
        self.parent.focus_set()
        self.destroy()


class TeamManager(tk.Toplevel):
    # ... (código inalterado)
    def __init__(self, parent, app_controller):
        super().__init__(parent)
        self.transient(parent)
        self.app_controller = app_controller
        self.title("Gerenciar Equipe")
        self.geometry("400x500")
        self.resizable(False, False)
        func_frame = ttk.LabelFrame(self, text="Equipe", padding=10)
        func_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        listbox_container = ttk.Frame(func_frame)
        listbox_container.pack(fill=tk.BOTH, expand=True)
        func_scrollbar = ttk.Scrollbar(listbox_container, orient="vertical")
        self.func_listbox = tk.Listbox(
            listbox_container,
            yscrollcommand=func_scrollbar.set,
            bg="white",
            borderwidth=0,
            highlightthickness=0,
            font=("Segoe UI", 10),
        )
        func_scrollbar.config(command=self.func_listbox.yview)
        func_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.func_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        btn_frame = ttk.Frame(func_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        btn_frame.columnconfigure(0, weight=1)
        btn_frame.columnconfigure(1, weight=1)
        ttk.Button(
            btn_frame, text="Adicionar", command=self.adicionar_funcionario
        ).grid(row=0, column=0, sticky="ew", padx=(0, 2))
        ttk.Button(btn_frame, text="Remover", command=self.remover_funcionario).grid(
            row=0, column=1, sticky="ew", padx=(2, 0)
        )
        self.populate_listbox()
        self.grab_set()

    def populate_listbox(self):
        self.func_listbox.delete(0, tk.END)
        for item in self.app_controller.get_funcionarios_para_display():
            self.func_listbox.insert(tk.END, item)

    def adicionar_funcionario(self):
        # Chama o novo método para obter a lista dinâmica de TODAS as tarefas ativas
        lista_de_tarefas = self.app_controller.get_tarefas_ativas()
        dialog = CustomDialog(
            self,
            lista_de_tarefas,  # Passa a nova lista dinâmica para o diálogo
        )
        if dialog.result:
            self.app_controller.adicionar_funcionario(*dialog.result)
            self.populate_listbox()

    def remover_funcionario(self):
        selecionados = self.func_listbox.curselection()
        if not selecionados:
            return

        display_string = self.func_listbox.get(selecionados[0])

        if messagebox.askyesno(
            "Confirmar Remoção",
            f"Remover '{display_string}' e todas as suas alocações?",
            parent=self,
        ):
            # --- INÍCIO DA CORREÇÃO ---
            # 1. Extrai apenas o nome do funcionário da string de exibição.
            try:
                nome_a_remover, _ = display_string.rsplit(" [", 1)
            except ValueError:
                # Fallback caso a string não tenha o formato esperado
                nome_a_remover = display_string

            # 2. Passa APENAS o nome para a função que limpa as alocações.
            self.app_controller.ui.remover_alocacoes_de_funcionario(nome_a_remover)
            # --- FIM DA CORREÇÃO ---

            # Esta parte permanece a mesma, mas agora usa a string original para remover da lista da equipe.
            self.app_controller.remover_funcionario(display_string)
            self.populate_listbox()


class SelecionarTarefaDialog(simpledialog.Dialog):
    """Diálogo para selecionar uma tarefa (disciplina/subcontrato) de uma lista."""

    def __init__(self, parent, title, tarefas_disponiveis):
        self.tarefas = tarefas_disponiveis
        self.resultado = None
        super().__init__(parent, title)

    def body(self, master):
        ttk.Label(master, text="Selecione a tarefa para adicionar:").pack(pady=5)
        self.combo = ttk.Combobox(
            master, values=self.tarefas, state="readonly", width=30
        )
        self.combo.pack(padx=10, pady=5)
        if self.tarefas:
            self.combo.current(0)
        return self.combo

    def apply(self):
        self.resultado = self.combo.get()


class GuiApp(tk.Tk):
    # ... (init e outros métodos iniciais sem alteração)
    def __init__(self, app_controller):
        super().__init__()
        self.app_controller = app_controller
        self.title("Planejador de Recursos por Lotes")
        self.geometry("1600x900")
        self.configure(bg="#F0F2F5")
        self.setup_styles()
        self.team_window = None

        self._is_formatting_date = False
        self._is_loading = False
        self.current_lote_name = None  # To control the currently visible lot.

        current_year = datetimeFootHolder.now().year
        footer_text = f"© {current_year} Canhedo Beppu Engenheiros Associados LTDA - Todos os direitos reservados."

        footer_frame = ttk.Frame(self, style="Card.TFrame", padding=(10, 5))
        footer_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(2, 0), padx=2)

        footer_label = ttk.Label(
            footer_frame,
            text=footer_text,
            font=("Segoe UI", 8),
            foreground="#6c757d",
        )
        footer_label.pack()

        main_frame = ttk.Frame(self, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.grid_rowconfigure(0, weight=1)
        main_frame.grid_columnconfigure(1, weight=1)
        control_panel = ttk.Frame(main_frame, width=300, style="Modern.TFrame")
        control_panel.grid(row=0, column=0, sticky="ns", padx=(0, 10))
        control_panel.grid_propagate(False)
        content_area = ttk.Frame(main_frame, style="Modern.TFrame")
        content_area.grid(row=0, column=1, sticky="nsew")
        content_area.grid_rowconfigure(0, weight=1)
        content_area.grid_columnconfigure(0, weight=1)
        self.create_control_panel_widgets(control_panel)
        self.notebook = ttk.Notebook(content_area)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # Consolidated Dashboard Tab (always present)
        dashboard_tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(dashboard_tab, text="Dashboard Consolidado")
        self.dash_tree = self._criar_treeview(dashboard_tab)

        # NEW: Main tab for managing lots (holds the combobox and inner notebook)
        self.lotes_main_tab = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(self.lotes_main_tab, text="Gerenciar Lotes")

        # Container for lot selection combobox
        lote_selector_frame = ttk.Frame(self.lotes_main_tab)
        lote_selector_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Label(lote_selector_frame, text="Selecione um Lote:").pack(
            side=tk.LEFT, padx=(0, 5)
        )
        self.lote_selection_combo = ttk.Combobox(
            lote_selector_frame, state="readonly", width=20
        )
        self.lote_selection_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.lote_selection_combo.bind("<<ComboboxSelected>>", self._on_lote_selected)

        # NEW: Inner Notebook for "Data Entry" and "Lot Dashboard" tabs PER LOT
        self.lote_inner_notebook = ttk.Notebook(self.lotes_main_tab)
        self.lote_inner_notebook.pack(fill=tk.BOTH, expand=True)

        # Frame that will contain the content of the selected lot (dynamic)
        # This will be packed into the "Data Entry" tab of the lote_inner_notebook
        self.current_lote_content_frame = ttk.Frame(
            self.lote_inner_notebook, style="Modern.TFrame"
        )
        # Frame for the individual lot dashboard
        self.current_lote_dashboard_frame = ttk.Frame(
            self.lote_inner_notebook, padding=5
        )
        self.current_lote_dash_tree = self._criar_treeview(
            self.current_lote_dashboard_frame
        )

        # This dictionary will store data AND widget references for the ACTIVE lot.
        # Other lots' data will be in memory, but their widgets destroyed when not active.
        self.lote_widgets = {}

        self.protocol("WM_DELETE_WINDOW", self._on_closing)

    def _on_lote_selected(self, event):
        """
        Saves the state of the currently visible lot (if any) and loads the state
        of the newly selected lot into the Combobox.
        """
        if self._is_loading:  # Prevents recursion during initial UI loading
            return

        selected_lote_name = self.lote_selection_combo.get()

        # 1. Save the state of the PREVIOUSLY visible lot (if any)
        if self.current_lote_name and self.current_lote_name != selected_lote_name:
            print(
                f"INFO: Saving state of lot '{self.current_lote_name}' before changing."
            )
            self._salvar_dados_lote_ui_para_memoria(self.current_lote_name)

        # 2. Update the currently selected lot
        self.current_lote_name = selected_lote_name

        # 3. Clear and configure the inner Notebook for the lot
        # Remove all tabs from the inner lot notebook to avoid duplicates
        for tab in self.lote_inner_notebook.tabs():
            self.lote_inner_notebook.forget(tab)

        # Re-add the "Data Entry" and "Lot Dashboard" tabs
        self.lote_inner_notebook.add(
            self.current_lote_content_frame, text="Entrada de Dados"
        )
        self.lote_inner_notebook.add(
            self.current_lote_dashboard_frame, text="Dashboard do Lote"
        )

        # 4. Clear the current lot content frame (data entry widgets)
        for widget in self.current_lote_content_frame.winfo_children():
            widget.destroy()

        # 5. Clear the current lot dashboard treeview
        self.current_lote_dash_tree.delete(*self.current_lote_dash_tree.get_children())

        # 6. Render the content of the newly selected lot
        if selected_lote_name:
            print(f"INFO: Loading and rendering lot '{selected_lote_name}'.")
            self._renderizar_lote_na_ui(selected_lote_name)

        # 7. After changing and rendering, recalculate to update dashboards
        self.processar_calculo(redirect_to_dashboard=False)

    def _salvar_dados_lote_ui_para_memoria(self, lote_nome):
        """
        Coleta os dados do lote visível na UI e os armazena no self.lote_widgets.
        Esta função é chamada APENAS para o lote que está visível e cujos widgets existem.
        """
        if lote_nome not in self.lote_widgets:
            return

        lote_data_in_memory = self.lote_widgets[lote_nome]

        # Check if the scrollable_frame exists and is visible before trying to save UI state
        if (
            lote_data_in_memory.get("scrollable_frame") is None
            or not lote_data_in_memory["scrollable_frame"].winfo_exists()
        ):
            # If the scrollable frame doesn't exist or isn't visible, there's no UI to save from.
            # This happens when switching away from a lot that was never rendered, or after a reset.
            print(
                f"DEBUG: Skipping UI data save for '{lote_nome}' - scrollable_frame not found or not exists."
            )
            return

        # NEW: Create a deep copy of the disciplines data from memory to work with,
        # ensuring we don't directly modify the live data structure until it's complete.
        # This is crucial because we'll be replacing the 'alocacoes' list entirely based on UI.
        disciplines_data_copy = {
            k: v.copy() for k, v in lote_data_in_memory["disciplinas"].items()
        }

        # Iterate over the active LabelFrames (tasks/disciplines) in the scrollable_frame of the active lot
        for child_frame in lote_data_in_memory["scrollable_frame"].winfo_children():
            if isinstance(child_frame, ttk.LabelFrame):  # Each task is a LabelFrame
                nome_tarefa = child_frame.cget("text")

                # Retrieve the corresponding discipline data from our copy
                disc_data_for_update = disciplines_data_copy.get(nome_tarefa)
                if not disc_data_for_update:
                    continue  # Should not happen if UI matches data structure

                start_date_entry = None
                end_date_entry = None

                # Find the date widgets within the task frame
                for widget in child_frame.winfo_children():
                    if isinstance(widget, ttk.Frame) and widget.winfo_children():
                        if any(
                            "Início" in c.cget("text")
                            for c in widget.winfo_children()
                            if hasattr(c, "cget")
                        ):
                            start_date_entry = widget.winfo_children()[
                                1
                            ]  # The Entry for Início
                            end_date_entry = widget.winfo_children()[
                                3
                            ]  # The Entry for Fim
                            break

                if not start_date_entry or not end_date_entry:
                    continue

                # Update cronograma dates based on current UI input
                disc_data_for_update["cronograma"] = {
                    "inicio": start_date_entry.get(),
                    "fim": end_date_entry.get(),
                }

                # CRITICAL FIX: Rebuild the 'alocacoes' list based on current UI state
                current_alocations_from_ui = []
                # Access disc_info directly from lote_data_in_memory as it contains the active widget refs.
                disc_info_for_active_widgets = lote_data_in_memory["disciplinas"].get(
                    nome_tarefa
                )

                if (
                    disc_info_for_active_widgets
                    and "alocacoes_widgets" in disc_info_for_active_widgets
                ):
                    for aloc_widget_dict in disc_info_for_active_widgets[
                        "alocacoes_widgets"
                    ]:
                        if aloc_widget_dict[
                            "frame"
                        ].winfo_exists():  # Only process if widget still exists
                            nome = aloc_widget_dict["combo_nome"].get()
                            cargo = aloc_widget_dict["combo_cargo"].get()
                            horas_str = (
                                aloc_widget_dict["entry"].get().replace(",", ".")
                            )

                            if nome and cargo and horas_str:
                                try:
                                    horas = float(horas_str)
                                    aloc_entry_data = {
                                        "funcionario": [nome, cargo],
                                        "horas_totais": horas,
                                    }
                                    current_alocations_from_ui.append(aloc_entry_data)
                                except ValueError:
                                    # Ignore allocations with invalid hours (ex: text)
                                    pass

                # Assign the newly collected allocations from UI to the discipline's data copy
                disc_data_for_update["alocacoes"] = current_alocations_from_ui

                # Update the original disciplines_data_copy with the modified disc_data_for_update
                disciplines_data_copy[nome_tarefa] = disc_data_for_update

        # Finally, update the lote_widgets in memory with the collected and updated data
        lote_data_in_memory["disciplinas"] = disciplines_data_copy
        print(f"INFO: Data for lot '{lote_nome}' saved from UI to memory.")

    def _renderizar_lote_na_ui(self, lote_nome):
        """
        Renders the content of a specific lot into the current_lote_content_frame.
        This function clears previous content and creates new widgets for the selected lot.
        """
        self._is_loading = True
        try:
            # Clear the current lot content frame before rendering new content
            for widget in self.current_lote_content_frame.winfo_children():
                widget.destroy()

            lote_data = self.lote_widgets.get(lote_nome)

            if not lote_data:
                messagebox.showerror(
                    "Erro", f"Dados para o lote '{lote_nome}' não encontrados."
                )
                return

            # Create the scrollable frame for the lot content
            canvas = tk.Canvas(
                self.current_lote_content_frame, highlightthickness=0, bg="#FFFFFF"
            )
            scrollbar = ttk.Scrollbar(
                self.current_lote_content_frame, orient="vertical", command=canvas.yview
            )
            scrollable_frame = ttk.Frame(canvas, style="Modern.TFrame")

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all")),
            )
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            def _on_mousewheel(event):
                canvas.yview_scroll(-1 * (event.delta // 120), "units")

            def _bind_scroll(event):
                canvas.bind_all("<MouseWheel>", _on_mousewheel)

            def _unbind_scroll(event):
                canvas.unbind_all("<MouseWheel>")

            canvas.bind("<Enter>", _bind_scroll)
            canvas.bind("<Leave>", _unbind_scroll)
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            # Store the reference to the scrollable_frame (used for saving data back)
            lote_data["scrollable_frame"] = (
                scrollable_frame  # Temporarily store for the active lot
            )

            # Recreate task frames and allocations based on lot data
            disciplinas_do_lote_data = lote_data.get("disciplinas", {})

            for (
                tarefa_nome,
                tarefa_data_from_memory,
            ) in disciplinas_do_lote_data.items():
                self._criar_frame_tarefa(
                    scrollable_frame, lote_nome, tarefa_nome, popular_defaults=False
                )

                tarefa_widgets = self.lote_widgets[lote_nome]["disciplinas"][
                    tarefa_nome
                ]

                # Populate date entries with saved values
                tarefa_widgets["start_date"].delete(0, tk.END)
                tarefa_widgets["start_date"].insert(
                    0, tarefa_data_from_memory.get("cronograma", {}).get("inicio", "")
                )
                tarefa_widgets["end_date"].delete(0, tk.END)
                tarefa_widgets["end_date"].insert(
                    0, tarefa_data_from_memory.get("cronograma", {}).get("fim", "")
                )

                # Clear allocation widgets list for the active task before populating
                tarefa_widgets["alocacoes_widgets"] = []

                # Populate allocations with saved data
                for aloc_data_from_memory in tarefa_data_from_memory.get(
                    "alocacoes", []
                ):
                    nome_func, cargo_func = aloc_data_from_memory.get(
                        "funcionario", (None, None)
                    )
                    horas = aloc_data_from_memory.get("horas_totais", "")

                    # Call adicionar_alocacao_equipe which creates the widgets and adds their refs
                    self.adicionar_alocacao_equipe(lote_nome, tarefa_nome)

                    if tarefa_widgets["alocacoes_widgets"]:
                        nova_aloc_widgets = tarefa_widgets["alocacoes_widgets"][-1]
                        combo_nome = nova_aloc_widgets["combo_nome"]
                        change_btn = nova_aloc_widgets["change_button"]

                        # Set the values directly (crucial for initial load)
                        combo_nome.set(nome_func if nome_func else "")
                        nova_aloc_widgets["combo_cargo"].set(
                            cargo_func if cargo_func else ""
                        )
                        if horas is not None:
                            nova_aloc_widgets["entry"].insert(
                                0, str(horas).replace(".", ",")
                            )

                        # Apply the locking logic based on the loaded value
                        if combo_nome.get():  # If a name was loaded, lock the combobox
                            combo_nome.config(state="disabled")
                            change_btn.config(state="normal")
                            combo_nome.config(values=[combo_nome.get(), ""])
                        else:  # If it's empty, leave it open for selection
                            combo_nome.config(state="readonly")
                            change_btn.config(state="disabled")
                            # And populate its initial options
                            self._populate_combobox_options(
                                combo_nome, nova_aloc_widgets["disciplina"]
                            )

            # Add the "+ Add Task" button at the end of the lot
            self._criar_botao_adicionar_tarefa(scrollable_frame, lote_nome)

        finally:
            self.update_idletasks()
            self._is_loading = False

    def _abrir_dialogo_adicionar_tarefa(self, lote_nome, parent_frame):
        """Abre um diálogo para o usuário selecionar uma tarefa ou criar uma nova."""
        todas_as_tarefas_add = self.app_controller.get_tarefas_adicionais_disponiveis()
        tarefas_ja_adicionadas = list(
            self.lote_widgets[lote_nome]["disciplinas"].keys()
        )

        tarefas_disponiveis = [
            t for t in todas_as_tarefas_add if t not in tarefas_ja_adicionadas
        ]

        if not tarefas_disponiveis:
            messagebox.showinfo(
                "Sem Tarefas",
                "Todas as tarefas adicionáveis já foram incluídas neste lote.",
                parent=self,
            )
            return

        dialog = SelecionarTarefaDialog(self, "Adicionar Tarefa", tarefas_disponiveis)
        tarefa_selecionada = dialog.resultado

        # Se o usuário cancelou a primeira janela, não faz nada
        if not tarefa_selecionada:
            return

        # --- INÍCIO DA NOVA LÓGICA ---
        nova_tarefa_final = None
        if tarefa_selecionada == "Outros...":
            # Se selecionou "Outros...", pede um nome customizado
            nome_customizado = simpledialog.askstring(
                "Nome da Tarefa", "Digite o nome para a nova tarefa:", parent=self
            )
            # Verifica se o usuário digitou algo e se o nome já não existe
            if nome_customizado and nome_customizado.strip():
                if nome_customizado in tarefas_ja_adicionadas:
                    messagebox.showwarning(
                        "Nome Duplicado",
                        "Uma tarefa com este nome já existe neste lote.",
                        parent=self,
                    )
                else:
                    nova_tarefa_final = nome_customizado.strip()
        else:
            # Se selecionou uma tarefa da lista, usa ela
            nova_tarefa_final = tarefa_selecionada
        # --- FIM DA NOVA LÓGICA ---

        # Cria o frame apenas se um nome de tarefa válido foi definido
        if nova_tarefa_final:
            # Ao adicionar uma nova tarefa, inicialize sua estrutura de dados no lote_widgets
            self.lote_widgets[lote_nome]["disciplinas"][nova_tarefa_final] = {
                "cronograma": {"inicio": "", "fim": ""},
                "alocacoes": [],  # Nova tarefa começa sem alocações
            }
            # E então chame _criar_frame_tarefa para construir a UI para ela
            self._criar_frame_tarefa(
                self.lote_widgets[lote_nome]["scrollable_frame"],
                lote_nome,
                nova_tarefa_final,
                popular_defaults=True,
            )

    def _on_data_keyrelease(self, event):
        """
        Chamado quando uma tecla é solta no campo de data.
        Formata o texto para DD/MM/AAAA e reposiciona o cursor corretamente.
        """
        entry = event.widget

        # Pega a posição do cursor e o texto ANTES da formatação
        cursor_pos = entry.index(tk.INSERT)
        texto_antes = entry.get()

        # Limpa o texto, mantendo apenas os dígitos
        digitos = "".join(filter(str.isdigit, texto_antes))
        digitos = digitos[:8]  # Limita a 8 dígitos (DDMMAAAA)

        # Formata o texto com as barras
        texto_formatado = ""
        if len(digitos) > 4:
            texto_formatado = f"{digitos[:2]}/{digitos[2:4]}/{digitos[4:]}"
        elif len(digitos) > 2:
            texto_formatado = f"{digitos[:2]}/{digitos[2:]}"
        else:
            texto_formatado = digitos

        # Calcula o ajuste do cursor
        # Se uma barra foi adicionada ANTES da posição original do cursor,
        # o cursor precisa ser movido para a direita.
        ajuste_cursor = 0
        if len(texto_formatado) > len(texto_antes):
            # Verifica se uma barra foi inserida nos primeiros 2 caracteres
            if (
                cursor_pos >= 2
                and texto_antes.count("/") == 0
                and texto_formatado.count("/") >= 1
            ):
                ajuste_cursor = 1
            # Verifica se uma barra foi inserida nos primeiros 5 caracteres
            if (
                cursor_pos >= 5
                and texto_antes.count("/") == 1
                and texto_formatado.count("/") >= 2
            ):
                ajuste_cursor = 1

        nova_pos_cursor = cursor_pos + ajuste_cursor

        # Atualiza o texto no Entry e reposiciona o cursor
        entry.delete(0, tk.END)
        entry.insert(0, texto_formatado)
        entry.icursor(nova_pos_cursor)

        return "break"  # Impede que outros bindings de tecla sejam executados

    def _on_combobox_click(self, event, aloc_widget):
        """
        Handler for when an allocation combobox is clicked.
        It forces an update of the combobox's options so the user can see available choices.
        """
        combobox_alvo = event.widget
        if self.current_lote_name:
            # When the user clicks, we force an update to show available options
            # so they can change the selection if desired.
            self._atualizar_valores_combobox_funcionario(
                combobox_alvo, aloc_widget["disciplina"], force_update=True
            )
            return

    def _atualizar_opcoes_funcionario(
        self,
        lote_nome,
        combobox_alvo,
        aloc_widget_clicado,
        force_update=True,  # New parameter to control when to update
    ):
        """
        Atualiza a lista de valores de uma combobox de funcionário.
        Se um funcionário já está selecionado e ainda é válido, a combobox é "travada"
        e não mostra outras opções, a menos que 'force_update' seja True.
        """
        if not combobox_alvo.winfo_exists():
            return

        selecao_atual = combobox_alvo.get()

        # Check if there's a valid selection and we are not forcing an update
        if (
            selecao_atual
            and selecao_atual in self.app_controller.get_todos_os_funcionarios()
            and not force_update
        ):
            # If a valid employee is already selected, 'lock' this combobox.
            # Its options will only be the current selection and an empty string.
            combobox_alvo.config(values=["", selecao_atual])
            combobox_alvo.set(selecao_atual)
            return

        # Proceed with normal filtering if no valid selection is made or force_update is True
        nomes_usados_no_lote = set()
        lote_data = self.lote_widgets.get(lote_nome, {})

        if "disciplinas" in lote_data:
            for disc_data_value in lote_data["disciplinas"].values():
                for aloc_widget in disc_data_value.get("alocacoes_widgets", []):
                    combo_iter = aloc_widget.get("combo_nome")

                    if (
                        combo_iter
                        and combo_iter.winfo_exists()
                        and combo_iter is not combobox_alvo
                    ):
                        selecao = combo_iter.get()
                        if selecao:
                            nomes_usados_no_lote.add(selecao)

        disciplina_alvo = aloc_widget_clicado.get("disciplina")
        nomes_base_para_tarefa = self.app_controller.get_funcionarios_para_tarefa(
            disciplina_alvo
        )

        opcoes_disponiveis = [
            nome for nome in nomes_base_para_tarefa if nome not in nomes_usados_no_lote
        ]

        # Always include the current selection if it's still valid (e.g., employee not removed)
        if (
            selecao_atual
            and selecao_atual in self.app_controller.get_todos_os_funcionarios()
            and selecao_atual not in opcoes_disponiveis
        ):
            opcoes_disponiveis.insert(0, selecao_atual)

        combobox_alvo.config(values=[""] + sorted(opcoes_disponiveis))

        # Restore the selection if it's still valid; otherwise, clear it.
        if selecao_atual in combobox_alvo["values"]:
            combobox_alvo.set(selecao_atual)
        else:
            combobox_alvo.set("")

    def limpar_alocacoes_de_funcionario_removido(self, nome_removido):
        """
        Iterates through all lots (visible and in memory), clearing combobox selections
        and removing the allocation from the persisted data structure.
        Also re-enables the comboboxes that had the removed employee.
        """
        for lote_name, lote_data in self.lote_widgets.items():
            if "disciplinas" in lote_data:
                for disc_name, disc_info in lote_data["disciplinas"].items():

                    updated_aloc_widgets = []
                    for aloc_item in disc_info.get("alocacoes_widgets", []):
                        if aloc_item["frame"].winfo_exists():
                            # Check if this allocation's combobox has the removed employee
                            if (
                                aloc_item.get("combo_nome")
                                and aloc_item["combo_nome"].get() == nome_removido
                            ):
                                aloc_item["combo_nome"].set("")  # Clear the selection
                                aloc_item["combo_nome"].config(
                                    state="readonly"
                                )  # Make it editable again
                                if aloc_item.get("change_button"):
                                    aloc_item["change_button"].config(
                                        state="disabled"
                                    )  # Disable change button
                                # IMPORTANT: Now update the options to allow selecting a new employee
                                self._populate_combobox_options(
                                    aloc_item["combo_nome"], aloc_item["disciplina"]
                                )
                                # Destroy the frame if you want to remove the UI row,
                                # otherwise just clear the combo and keep the row.
                                # Given previous intent, it seems like we destroy the UI row.
                                aloc_item["frame"].destroy()
                            else:
                                updated_aloc_widgets.append(aloc_item)
                        pass  # If frame doesn't exist, it's already removed or not rendered.
                    disc_info["alocacoes_widgets"] = updated_aloc_widgets

                    new_aloc_data_list = []
                    for aloc_data in disc_info.get("alocacoes", []):
                        if (
                            aloc_data.get("funcionario")
                            and aloc_data["funcionario"][0] == nome_removido
                        ):
                            pass  # Remove this allocation from persisted data
                        else:
                            new_aloc_data_list.append(aloc_data)
                    disc_info["alocacoes"] = new_aloc_data_list

        # After cleaning up all allocations, trigger a general recalculation
        # and re-render the current lot to ensure UI consistency.
        if self.current_lote_name and self.current_lote_name in self.lote_widgets:
            self._renderizar_lote_na_ui(
                self.current_lote_name
            )  # This will rebuild the lot UI

        self.processar_calculo(redirect_to_dashboard=False)

    def _on_horas_keyrelease(self, event):
        entry = event.widget
        texto_atual = entry.get()
        texto_limpo = ""
        virgula_encontrada = False
        for char in texto_atual:
            if char.isdigit():
                texto_limpo += char
            elif char in (",", ".") and not virgula_encontrada:
                texto_limpo += ","
                virgula_encontrada = True

        if texto_limpo != texto_atual:
            cursor_pos = entry.index(tk.INSERT)
            entry.delete(0, tk.END)
            entry.insert(0, texto_limpo)
            if cursor_pos > 0:
                entry.icursor(cursor_pos - 1)

    def _on_closing(self):
        # Se o MAC não for autorizado, apenas fecha a janela sem salvar
        if not self.app_controller.mac_autorizado:
            self.destroy()
            return

        # Se for autorizado, salva o estado e fecha
        # Salva o estado do lote atualmente visível antes de fechar
        if self.current_lote_name:
            self._salvar_dados_lote_ui_para_memoria(self.current_lote_name)

        self.salvar_estado()
        self.destroy()

    def create_control_panel_widgets(self, parent):
        # Frame da Equipe
        team_frame = ttk.LabelFrame(parent, text="Equipe", padding=10)
        team_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=(10, 5))
        team_frame.columnconfigure(0, weight=1)

        self.btn_team = ttk.Button(
            team_frame,
            text="Gerenciar Equipe",
            command=self.open_team_manager,
        )
        self.btn_team.grid(row=0, column=0, sticky="ew", ipady=4)

        # Frame de Configuração
        setup_frame = ttk.LabelFrame(parent, text="Configuração", padding=10)
        setup_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)

        ttk.Label(setup_frame, text="Horas Trabalháveis/Mês:").pack(
            anchor=tk.W, pady=(0, 2)
        )
        self.horas_mes_entry = ttk.Entry(setup_frame)
        self.horas_mes_entry.insert(0, "160")
        self.horas_mes_entry.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(setup_frame, text="Número de Lotes:").pack(anchor=tk.W, pady=(0, 2))
        self.num_lotes_entry = ttk.Entry(setup_frame)
        self.num_lotes_entry.insert(0, "1")
        self.num_lotes_entry.pack(fill=tk.X, pady=(0, 10))
        self.btn_lotes = ttk.Button(
            setup_frame,
            text="Gerar Lotes",
            command=self.gerar_abas_lotes,
        )

        self.btn_lotes.pack(fill=tk.X)

        # Frame de Ações
        action_frame = ttk.LabelFrame(parent, text="Ações", padding=10)
        action_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        action_frame.columnconfigure(0, weight=1)

        self.btn_calcular = ttk.Button(
            action_frame,
            text="Calcular Alocação",
            command=lambda: self.processar_calculo(redirect_to_dashboard=True),
            style="Accent.TButton",
        )
        self.btn_calcular.grid(row=0, column=0, sticky="ew", ipady=6, pady=2)

        self.btn_exportar_cronograma = ttk.Button(
            action_frame,
            text="Exportar Relatório de Cronograma",
            command=self.exportar_para_excel,
            style="Accent.TButton",
        )
        self.btn_exportar_cronograma.grid(row=1, column=0, sticky="ew", ipady=6, pady=2)

        self.btn_exportar_quantidades = ttk.Button(
            action_frame,
            text="Exportar Planilha de Quantidades",
            command=self.exportar_resumo_equipe,
            style="Accent.TButton",
        )
        self.btn_exportar_quantidades.grid(
            row=2, column=0, sticky="ew", ipady=6, pady=2
        )

        ttk.Separator(action_frame, orient="horizontal").grid(
            row=3, column=0, sticky="ew", pady=(10, 5)
        )

        self.btn_salvar = ttk.Button(
            action_frame,
            text="Salvar Dados",
            command=self._salvar_dados_manualmente,
            style="TButton",
        )
        self.btn_salvar.grid(row=4, column=0, sticky="ew", ipady=6, pady=2)

        # MODIFIED: Placed "Resetar Tudo" button AFTER "Salvar Dados"
        self.btn_resetar = ttk.Button(
            action_frame,
            text="Resetar Tudo",
            command=self.resetar_interface_completa,  # Now calls the confirmation method
        )
        self.btn_resetar.grid(row=5, column=0, sticky="ew", ipady=4, pady=2)

    def _salvar_dados_manualmente(self):
        """
        Salva o estado atual da UI para a memória e depois para o arquivo JSON.
        """
        try:
            # Primeiro, garanta que os dados do lote atualmente visível na UI
            # sejam salvos na estrutura de dados em memória (self.lote_widgets).
            if self.current_lote_name:
                self._salvar_dados_lote_ui_para_memoria(self.current_lote_name)

            # Agora, chame a função principal de salvamento que escreve para o JSON.
            self.salvar_estado()
            messagebox.showinfo("Salvo", "Dados salvos com sucesso!", parent=self)
        except Exception as e:
            messagebox.showerror(
                "Erro ao Salvar",
                f"Ocorreu um erro ao salvar os dados: {e}",
                parent=self,
            )

    def resetar_interface_completa(self):
        """
        Asks for confirmation and, if positive, resets the entire application to its initial state.
        """
        if not messagebox.askyesno(
            "Confirmar Reset",
            "Are you sure you want to delete ALL data and restart the interface?\n\nThis action cannot be undone.",
            parent=self,
        ):
            return

        print("INFO: Initiating complete application reset.")

        self.app_controller.resetar_equipe()

        self.horas_mes_entry.delete(0, tk.END)
        self.horas_mes_entry.insert(0, "160")
        self.num_lotes_entry.delete(0, tk.END)
        self.num_lotes_entry.insert(0, "1")

        self.lote_selection_combo.set("")
        self.lote_selection_combo.config(values=[])

        # Destroy all cached UI frames for each lot
        for lote_name, lote_data in self.lote_widgets.items():
            if (
                lote_data.get("data_entry_tab_frame")
                and lote_data["data_entry_tab_frame"].winfo_exists()
            ):
                lote_data["data_entry_tab_frame"].destroy()
            if (
                lote_data.get("dashboard_tab_frame")
                and lote_data["dashboard_tab_frame"].winfo_exists()
            ):
                lote_data["dashboard_tab_frame"].destroy()

        # Clear the inner notebook tabs
        for tab in self.lote_inner_notebook.tabs():
            self.lote_inner_notebook.forget(tab)

        # Reset current lot tracking
        self.current_lote_name = None

        # Clear all lot data from memory
        self.lote_widgets.clear()

        # Recreate default lots (which will also populate the combo and trigger rendering of the first lot)
        self.gerar_abas_lotes(popular_defaults=True)

        # Clear the consolidated dashboard table
        self.dash_tree.delete(*self.dash_tree.get_children())

        # Clear old report data
        self.ultimo_df_consolidado = None
        if self.app_controller.ultimo_dashboards_lotes:
            self.app_controller.ultimo_dashboards_lotes.clear()

        # The individual lot dashboard tree is now handled by the _renderizar_lote_na_ui for the active lot.
        # So, no need to explicitly clear current_lote_dash_tree here, as it will be cleared by the next render.
        # self.current_lote_dash_tree.delete(*self.current_lote_dash_tree.get_children())

        messagebox.showinfo(
            "Reset Concluded",
            "The application has been reset to its initial state.",
            parent=self,
        )

    def set_modo_restringido(self, restringido):
        """Habilita ou desabilita as funcionalidades principais da aplicação."""
        if restringido:
            novo_estado = tk.DISABLED
            # Mostra um aviso claro para o usuário
            messagebox.showerror(
                "Licença Inválida",
                "Este aplicativo não está licenciado para esta máquina.\nAs funcionalidades de cálculo, exportação e salvamento foram desabilitadas.",
            )
        else:
            novo_estado = tk.NORMAL

        # Altera o estado dos botões
        if hasattr(self, "btn_calcular"):  # Verifica se os botões já existem
            self.btn_team.config(state=novo_estado)
            self.btn_lotes.config(state=novo_estado)
            self.btn_calcular.config(state=novo_estado)
            self.btn_exportar_cronograma.config(state=novo_estado)
            self.btn_exportar_quantidades.config(state=novo_estado)

    # NOVO: Função para exportar o resumo da equipe.
    def exportar_resumo_equipe(self):
        try:
            # Get the summary data grouped by lot
            # Assuming your app_controller has a method to get this, e.g., get_resumo_equipe_por_lote()
            # This method should return a dictionary where keys are lot names and values are DataFrames
            # Or you might need to iterate through self.lote_widgets and construct these DFs.
            lotes_resumo_dfs = (
                self.app_controller.get_resumo_equipe_por_lote()
            )  # THIS IS A NEW METHOD IN APP_CONTROLLER

            if not lotes_resumo_dfs:
                messagebox.showwarning(
                    "Aviso",
                    "Não há dados de alocação da equipe para exportar.",
                    parent=self,
                )
                return

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Arquivo Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")],
                title="Salvar Resumo da Equipe por Lote",
                initialfile=f"Resumo_Equipe_Lotes_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx",
            )
            if not filepath:
                return

            cargos = self.app_controller.get_cargos_disponiveis()

            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                book = writer.book  # Get the workbook object early

                # --- DEFINE AND ADD NAMED STYLES ONCE HERE ---
                input_style = NamedStyle(name="input_style_resumo")
                input_style.fill = PatternFill(
                    start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
                )
                input_style.number_format = '"R$" #,##0.00'
                # Note: data_font is likely defined within _write_styled_resumo_to_excel.
                # You might need to pass data_font or define common fonts here too.
                # For simplicity, let's assume data_font is accessible or passed.
                # For this example, let's just make sure input_style has a font.
                input_style.font = Font(
                    name="Arial", size=10
                )  # Or pass 'data_font' if available here

                if input_style.name not in book.style_names:
                    book.add_named_style(input_style)
                # --- END OF NAMED STYLE DEFINITION ---

                # Write each lot's summary to a separate sheet
                for lote_name, df_resumo_lote in lotes_resumo_dfs.items():
                    if not df_resumo_lote.empty:
                        sheet_name = f"Lote {lote_name}"
                        self._write_styled_resumo_to_excel(
                            writer, sheet_name, df_resumo_lote, cargos
                        )

                # Remove the default sheet if it exists
                if "Sheet" in writer.book.sheetnames:
                    writer.book.remove(writer.book["Sheet"])

            messagebox.showinfo(
                "Sucesso",
                f"Resumo da equipe exportado com sucesso para:\n{filepath}",
                parent=self,
            )

        except Exception as e:
            messagebox.showerror(
                "Erro na Exportação",
                f"Ocorreu um erro ao exportar o resumo da equipe:\n{e}",
                parent=self,
            )

    # NOVO: Método para escrever o resumo da equipe com formatação avançada.
    def _write_styled_resumo_to_excel(self, writer, sheet_name, df, cargos):
        book = writer.book
        ws = book.create_sheet(title=sheet_name)
        writer.sheets[sheet_name] = ws
        ws.sheet_view.showGridLines = False

        # --- ESTILOS PADRONIZADOS ---
        # (styles definitions remain the same)
        header_font = Font(name="Arial", bold=True, color="000000", size=11)
        header_fill = PatternFill(
            start_color="E68A00", end_color="E68A00", fill_type="solid"
        )
        category_font = Font(name="Arial", bold=True, size=12)
        category_fill = PatternFill(
            start_color="E6AF60", end_color="E6AF60", fill_type="solid"
        )
        data_font = Font(name="Arial", size=10)
        thin_side = Side(border_style="thin", color="BFBFBF")
        full_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )
        horas_fill = PatternFill(
            start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"
        )  # Amarelo claro

        # --- 1. TABELA DE PREÇOS ---
        ws.cell(row=2, column=2, value="Tabela de Preços por Cargo").font = Font(
            bold=True, size=14
        )
        price_header_font = Font(name="Arial", bold=True, size=11)
        cell_cargo_header = ws.cell(row=3, column=2, value="Cargo")
        cell_cargo_header.font = price_header_font
        cell_cargo_header.border = full_border
        cell_preco_header = ws.cell(row=3, column=3, value="Preço/Hora (R$)")
        cell_preco_header.font = price_header_font
        cell_preco_header.border = full_border

        start_row_price_table = 4
        for i, cargo_nome in enumerate(cargos, start=start_row_price_table):
            ws.cell(row=i, column=2, value=cargo_nome).font = data_font
            price_cell = ws.cell(row=i, column=3)
            price_cell.style = "input_style_resumo"

        end_row_price_table = start_row_price_table + len(cargos) - 1

        # MODIFICATION START: Make named range name unique per sheet
        unique_price_table_name = f"TabelaPrecos_{sheet_name.replace(' ', '_')}"  # e.g., "TabelaPrecos_Lote_1"
        price_table_range_str = f"'{sheet_name}'!${openpyxl.utils.get_column_letter(2)}${start_row_price_table}:${openpyxl.utils.get_column_letter(3)}${end_row_price_table}"

        # Get the sheet index for local scope
        sheet_index = book.sheetnames.index(sheet_name)

        named_range_obj = openpyxl.workbook.defined_name.DefinedName(
            unique_price_table_name,  # Use the unique name
            localSheetId=sheet_index,
            attr_text=price_table_range_str,
        )
        book.defined_names[named_range_obj.name] = named_range_obj

        # MODIFICATION END

        # --- 2. TABELA PRINCIPAL DE DADOS ---
        df["Cargo"] = pd.Categorical(df["Cargo"], categories=cargos, ordered=True)
        df_sorted = df.sort_values(by=["Disciplina", "Cargo", "Funcionário"])
        start_row_main_table = end_row_price_table + 3

        headers = [
            "Funcionário",
            "Cargo",
            "Disciplina",
            "Horas Totais Alocadas",
            "Valor Total (R$)",
        ]
        for c_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=start_row_main_table, column=c_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill

        rows = dataframe_to_rows(df_sorted, index=False, header=False)
        current_row_excel = start_row_main_table + 1
        last_discipline = None

        for row_data_tuple in rows:
            disciplina_atual = row_data_tuple[2]
            if disciplina_atual != last_discipline:
                cat_cell = ws.cell(
                    row=current_row_excel, column=1, value=disciplina_atual
                )
                cat_cell.font = category_font
                cat_cell.fill = category_fill
                ws.merge_cells(
                    start_row=current_row_excel,
                    start_column=1,
                    end_row=current_row_excel,
                    end_column=len(headers),
                )
                last_discipline = disciplina_atual
                current_row_excel += 1

            for c_idx, value in enumerate(row_data_tuple, 1):
                cell = ws.cell(row=current_row_excel, column=c_idx, value=value)
                if c_idx == 4:  # Coluna 4 é "Horas Totais Alocadas"
                    cell.fill = horas_fill
                    cell.font = data_font

            cargo_cell_coord = (
                f"{openpyxl.utils.get_column_letter(2)}{current_row_excel}"
            )
            horas_cell_coord = (
                f"{openpyxl.utils.get_column_letter(4)}{current_row_excel}"
            )
            # MODIFICATION START: Use the unique named range in the formula
            formula = f'=IFERROR(VLOOKUP({cargo_cell_coord},{unique_price_table_name},2,FALSE)*{horas_cell_coord},"-")'
            # MODIFICATION END

            valor_cell = ws.cell(row=current_row_excel, column=5, value=formula)
            valor_cell.number_format = '"R$" #,##0.00'
            current_row_excel += 1
        end_row_main_table = current_row_excel - 1

        # --- 3. SEÇÃO DE TOTAIS E RESUMO FINANCEIRO ---
        total_geral_row = end_row_main_table + 1
        valor_col_letter = openpyxl.utils.get_column_letter(5)
        total_geral_label_cell = ws.cell(
            row=total_geral_row, column=4, value="Total Geral"
        )
        total_geral_label_cell.font = Font(name="Arial", bold=True, size=11)
        total_geral_label_cell.alignment = Alignment(horizontal="right")

        total_geral_valor_cell = ws.cell(
            row=total_geral_row,
            column=5,
            value=f"=SUM({valor_col_letter}${start_row_main_table+1}:${valor_col_letter}{end_row_main_table})",
        )
        total_geral_valor_cell.font = Font(name="Arial", bold=True, size=11)
        total_geral_valor_cell.number_format = '"R$" #,##0.00'

        start_row_summary = total_geral_row + 3
        ws.cell(
            row=start_row_summary, column=2, value="Resumo Financeiro por Disciplina"
        ).font = Font(bold=True, size=14)
        summary_headers = [
            "Disciplina",
            "Total HH (R$)",
            "Total por Desenho (R$)",
            "Subtotal (R$)",
        ]
        for c_idx, text in enumerate(summary_headers, 2):
            cell = ws.cell(row=start_row_summary + 1, column=c_idx, value=text)
            cell.font = header_font
            cell.fill = header_fill

        disciplina_col_main_table = openpyxl.utils.get_column_letter(3)
        valor_col_main_table = openpyxl.utils.get_column_letter(5)
        disciplinas_unicas = df["Disciplina"].unique()

        start_row_summary_data = start_row_summary + 2
        for i, disciplina_nome in enumerate(disciplinas_unicas):
            current_summary_row = start_row_summary_data + i

            ws.cell(row=current_summary_row, column=2, value=disciplina_nome)

            criteria_cell = ws.cell(row=current_summary_row, column=2).coordinate
            formula_sumif = f"=SUMIF({disciplina_col_main_table}${start_row_main_table+1}:{disciplina_col_main_table}${end_row_main_table}, {criteria_cell}, {valor_col_main_table}${start_row_main_table+1}:{valor_col_main_table}${end_row_main_table})"
            total_cell = ws.cell(row=current_summary_row, column=3, value=formula_sumif)
            total_cell.number_format = '"R$" #,##0.00'

            desconto_cell = ws.cell(row=current_summary_row, column=4)
            desconto_cell.style = "input_style_resumo"

            formula_subtotal = f"={ws.cell(row=current_summary_row, column=3).coordinate}-{ws.cell(row=current_summary_row, column=4).coordinate}"
            subtotal_cell = ws.cell(
                row=current_summary_row, column=5, value=formula_subtotal
            )
            subtotal_cell.number_format = '"R$" #,##0.00'

        # --- CORREÇÃO DA POSIÇÃO DO TOTAL FINAL ---
        last_data_row_summary = start_row_summary_data + len(disciplinas_unicas) - 1
        total_final_row = last_data_row_summary + 1

        total_final_label = ws.cell(row=total_final_row, column=4, value="Total Final")
        total_final_label.font = Font(name="Arial", bold=True, size=12)
        total_final_label.alignment = Alignment(horizontal="right")

        subtotal_col_letter = openpyxl.utils.get_column_letter(5)
        # CORREÇÃO DA FÓRMULA DE SOMA
        formula_total_final = f"=SUM({subtotal_col_letter}{start_row_summary_data}:{subtotal_col_letter}{last_data_row_summary})"
        total_final_valor = ws.cell(
            row=total_final_row, column=5, value=formula_total_final
        )
        total_final_valor.font = Font(name="Arial", bold=True, size=12)
        total_final_valor.number_format = '"R$" #,##0.00'

        # --- 4. FORMATAÇÃO FINAL ---
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value is not None or cell.style == "input_style_resumo":
                    cell.border = full_border

        for i in range(1, len(headers) + 2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 25

    def open_team_manager(self):
        if self.team_window and self.team_window.winfo_exists():
            self.team_window.focus()
        else:
            self.team_window = TeamManager(self, self.app_controller)

    # ... (setup_styles inalterado)
    def setup_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        BG_COLOR, PRIMARY_COLOR, TEXT_COLOR, LIGHT_TEXT_COLOR, BORDER_COLOR = (
            "#F0F2F5",
            "#0078D7",
            "#212121",
            "#616161",
            "#CCCCCC",
        )
        style.configure(
            ".", background=BG_COLOR, foreground=TEXT_COLOR, font=("Segoe UI", 10)
        )
        style.configure("TFrame", background=BG_COLOR)
        style.configure(
            "Modern.TFrame",
            background="white",
            relief="solid",
            borderwidth=1,
            bordercolor=BORDER_COLOR,
        )
        style.configure("TLabel", background=BG_COLOR)
        style.configure("TLabelframe", background=BG_COLOR, bordercolor=BORDER_COLOR)
        style.configure(
            "TLabelframe.Label", background=BG_COLOR, font=("Segoe UI", 10, "bold")
        )
        style.configure(
            "TButton",
            padding=6,
            relief="flat",
            background="#E1E1E1",
            font=("Segoe UI", 10),
        )
        style.map("TButton", background=[("active", "#CFCFCF")])
        style.configure(
            "Accent.TButton",
            foreground="white",
            background=PRIMARY_COLOR,
            font=("Segoe UI", 10, "bold"),
        )
        style.map("Accent.TButton", background=[("active", "#005A9E")])
        style.configure("TNotebook", background=BG_COLOR, borderwidth=0)
        style.configure("TNotebook.Tab", padding=[10, 5], font=("Segoe UI", 10))
        try:
            style.layout(
                "TNotebook.Tab",
                [
                    (
                        "Notebook.tab",
                        {
                            "sticky": "nswe",
                            "children": [
                                (
                                    "Notebook.padding",
                                    {
                                        "side": "top",
                                        "sticky": "nswe",
                                        "children": [
                                            (
                                                "Notebook.label",
                                                {"side": "top", "sticky": ""},
                                            )
                                        ],
                                    },
                                )
                            ],
                        },
                    )
                ],
            )
        except tk.TclError:
            pass
        style.map(
            "TNotebook.Tab",
            background=[("selected", "white"), ("!selected", BG_COLOR)],
            foreground=[("selected", PRIMARY_COLOR), ("!selected", LIGHT_TEXT_COLOR)],
        )
        style.configure(
            "Modern.Treeview", background="white", fieldbackground="white", rowheight=25
        )
        style.configure(
            "Modern.Treeview.Heading", font=("Segoe UI", 10, "bold"), padding=5
        )
        style.map("Modern.Treeview.Heading", background=[("active", "#E5E5E5")])

    def remover_alocacoes_de_funcionario(self, nome_removido):
        for lote_nome, lote_data in self.lote_widgets.items():
            if "disciplinas" in lote_data:
                for disc_name, disc_info in lote_data["disciplinas"].items():

                    updated_aloc_widgets = []
                    for aloc_item in disc_info.get("alocacoes_widgets", []):
                        if aloc_item["frame"].winfo_exists():
                            if (
                                aloc_item.get("combo_nome")
                                and aloc_item["combo_nome"].get() == nome_removido
                            ):
                                aloc_item["frame"].destroy()
                            else:
                                updated_aloc_widgets.append(aloc_item)
                        pass
                    disc_info["alocacoes_widgets"] = updated_aloc_widgets

                    new_aloc_data_list = []
                    for aloc_data in disc_info.get("alocacoes", []):
                        if (
                            aloc_data.get("funcionario")
                            and aloc_data["funcionario"][0] == nome_removido
                        ):
                            pass
                        else:
                            new_aloc_data_list.append(aloc_data)
                    disc_info["alocacoes"] = new_aloc_data_list

        if self.current_lote_name and self.current_lote_name in self.lote_widgets:
            # Re-render the current lot to reflect the removal
            self._renderizar_lote_na_ui(self.current_lote_name)

        self.processar_calculo(redirect_to_dashboard=False)

    def _criar_treeview(self, parent_frame):
        # ... (código inalterado)
        container = ttk.Frame(parent_frame)
        container.pack(fill=tk.BOTH, expand=True)
        tree = ttk.Treeview(container, style="Modern.Treeview")
        yscroll = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(container, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        xscroll.pack(side=tk.BOTTOM, fill=tk.X)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        return tree

    # MODIFICADO: Atualiza todas as listas de forma mais inteligente.
    def atualizar_lista_funcionarios(self):
        if self.team_window and self.team_window.winfo_exists():
            self.team_window.populate_listbox()

        # Iterate over all lots to ensure all comboboxes are updated with the latest employee list and availability.
        # This is particularly important when employees are added/removed.
        for lote_name, lote_data in self.lote_widgets.items():
            if "disciplinas" in lote_data:
                for nome_da_tarefa, disc_widgets in lote_data["disciplinas"].items():
                    for aloc_widget in disc_widgets.get("alocacoes_widgets", []):
                        combo_nome = aloc_widget.get("combo_nome")
                        change_btn = aloc_widget.get(
                            "change_button"
                        )  # Get reference to new button

                        if combo_nome and combo_nome.winfo_exists():
                            # If the combobox currently has a selected value and that employee is no longer registered,
                            # clear the combobox and enable it for new selection.
                            current_selection = combo_nome.get()
                            if (
                                current_selection
                                and current_selection
                                not in self.app_controller.get_todos_os_funcionarios()
                            ):
                                combo_nome.set("")  # Clear invalid selection
                                combo_nome.config(state="readonly")  # Make it editable
                                if change_btn:
                                    change_btn.config(
                                        state="disabled"
                                    )  # Disable change button if no selection
                                # Repopulate options after clearing
                                self._populate_combobox_options(
                                    combo_nome, aloc_widget["disciplina"]
                                )
                            elif (
                                not current_selection
                            ):  # If it's empty, populate it for new selection
                                combo_nome.config(
                                    state="readonly"
                                )  # Ensure it's editable
                                if change_btn:
                                    change_btn.config(
                                        state="disabled"
                                    )  # Disable change button if no selection
                                self._populate_combobox_options(
                                    combo_nome, aloc_widget["disciplina"]
                                )
                            else:  # If it has a valid selection, ensure it stays disabled/locked
                                combo_nome.config(state="disabled")
                                if change_btn:
                                    change_btn.config(state="normal")
                                combo_nome.config(
                                    values=[current_selection, ""]
                                )  # Reinforce locked values

    def _atualizar_opcoes_funcionarios(
        self, lote_nome, disciplina_alvo, combobox_clicada
    ):
        """
        Filtra a lista de um Combobox para mostrar apenas funcionários não utilizados no lote.
        """
        nomes_ja_selecionados = set()
        lote_data = self.lote_widgets.get(lote_nome, {})

        # 1. Coleta todos os nomes já selecionados em OUTROS comboboxes do mesmo lote
        if "disciplinas" in lote_data:
            for disc_widgets in lote_data["disciplinas"].values():
                for aloc_widget in disc_widgets.get("alocacoes_widgets", []):
                    combo_atual = aloc_widget.get("combo_nome")

                    # Verifica se o widget ainda existe e se NÃO é o que foi clicado
                    if (
                        combo_atual
                        and combo_atual.winfo_exists()
                        and combo_atual is not combobox_clicada
                    ):
                        selecao = combo_atual.get()
                        if selecao:  # Garante que não adicione strings vazias
                            nomes_ja_selecionados.add(selecao)

        # 2. Pega a lista base de funcionários para a disciplina
        nomes_base = self.app_controller.get_nomes_funcionarios_por_disciplina(
            disciplina_alvo
        )

        # 3. Filtra a lista, removendo os nomes já selecionados
        opcoes_disponiveis = [
            nome for nome in nomes_base if nome not in nomes_ja_selecionados
        ]

        # 4. Garante que a seleção atual do combobox clicado esteja na lista (o "edge case")
        selecao_atual = combobox_clicada.get()
        if selecao_atual and selecao_atual not in opcoes_disponiveis:
            opcoes_disponiveis.insert(0, selecao_atual)

        # 5. Atualiza os valores do combobox, sempre incluindo a opção vazia
        combobox_clicada["values"] = [""] + sorted(opcoes_disponiveis)

    # ### INÍCIO DA MODIFICAÇÃO: Helper para remover linha de alocação ###
    def _remover_linha_alocacao(self, lote_nome, disc_nome, widget_dict_a_remover):
        """
        Remove uma linha de alocação da UI e da estrutura de dados persistida.
        """
        if widget_dict_a_remover["frame"].winfo_exists():
            widget_dict_a_remover["frame"].destroy()  # Destrói o widget da UI

        try:
            # Remove a referência do widget da lista de widgets ativos do lote
            lista_widgets_ativos = self.lote_widgets[lote_nome]["disciplinas"][
                disc_nome
            ]["alocacoes_widgets"]
            if widget_dict_a_remover in lista_widgets_ativos:
                lista_widgets_ativos.remove(widget_dict_a_remover)

            # Remove a alocação dos dados persistidos (identificando-a pelos valores)
            alocacoes_data = self.lote_widgets[lote_nome]["disciplinas"][disc_nome][
                "alocacoes"
            ]

            # Pega os valores da alocação que está sendo removida da UI
            nome_removido_ui = widget_dict_a_remover["combo_nome"].get()
            cargo_removido_ui = widget_dict_a_remover["combo_cargo"].get()
            horas_removidas_str_ui = (
                widget_dict_a_remover["entry"].get().replace(",", ".")
            )
            horas_removidas_ui = (
                float(horas_removidas_str_ui) if horas_removidas_str_ui else 0.0
            )

            found_and_removed = False
            for i, aloc_entry_data in enumerate(alocacoes_data):
                func = aloc_entry_data.get("funcionario")
                horas = aloc_entry_data.get("horas_totais")

                # Compara com base nos valores da UI. Usa tolerância para floats.
                if (
                    func
                    and func[0] == nome_removido_ui
                    and func[1] == cargo_removido_ui
                    and abs(horas - horas_removidas_ui) < 0.001
                ):

                    alocacoes_data.pop(i)  # Remove do dicionário de dados persistidos
                    found_and_removed = True
                    break

            # Após remover, recalcula
            self.processar_calculo(redirect_to_dashboard=False)

        except (KeyError, ValueError, IndexError) as e:
            print(f"Erro ao remover linha de alocação: {e}")
            pass

    # ### FIM DA MODIFICAÇÃO ###

    def _enable_combobox_for_change(self, lote_nome, disciplina, aloc_widget_dict):
        """
        Enables the employee combobox for editing and populates its options.
        Disables the "Change" button while in edit mode.
        """
        combo_nome = aloc_widget_dict["combo_nome"]
        change_btn = aloc_widget_dict["change_button"]

        # If it's already enabled, perhaps treat as a cancel or re-lock
        if (
            combo_nome["state"] == "readonly"
        ):  # It's currently editable (or just became editable)
            # If it has a value, and that value is still valid, lock it again
            if (
                combo_nome.get()
                and combo_nome.get() in self.app_controller.get_todos_os_funcionarios()
            ):
                combo_nome.config(state="disabled")
                change_btn.config(state="normal")
                # Ensure values are just the selected one and empty for a locked state
                combo_nome.config(values=[combo_nome.get(), ""])
                self.processar_calculo(
                    redirect_to_dashboard=False
                )  # Recalculate if a selection was made
            else:
                # If it's empty or invalid, keep it open but update options
                self._populate_combobox_options(combo_nome, disciplina)
                change_btn.config(state="disabled")  # No selection to "change" yet

        else:  # It's currently disabled (locked), so enable it
            combo_nome.config(state="readonly")  # Make it editable
            change_btn.config(
                state="disabled"
            )  # Temporarily disable change button during edit
            self._populate_combobox_options(
                combo_nome, disciplina
            )  # Repopulate options
            combo_nome.focus_set()  # Give focus to the combobox

    # MODIFICADO: A combobox de alocação agora usa a lista filtrada.
    def adicionar_alocacao_equipe(
        self, lote_nome, disciplina, cargo_a_definir=None, horas_a_definir=None
    ):
        disc_widgets = self.lote_widgets[lote_nome]["disciplinas"][disciplina]
        frame = disc_widgets["alocacoes_frame"]

        row_frame = ttk.Frame(frame)
        row_frame.pack(fill=tk.X, pady=2)

        ttk.Label(row_frame, text="Funcionário:").pack(side=tk.LEFT, padx=(0, 5))
        combo_nome = ttk.Combobox(
            row_frame, state="readonly", width=25
        )  # Initial state
        combo_nome.pack(side=tk.LEFT, padx=(0, 10))

        ttk.Label(row_frame, text="Cargo:").pack(side=tk.LEFT, padx=(0, 5))
        cargos_disponiveis = self.app_controller.get_cargos_disponiveis()
        combo_cargo = ttk.Combobox(
            row_frame, values=cargos_disponiveis, state="readonly", width=20
        )
        combo_cargo.pack(side=tk.LEFT, padx=(0, 10))
        if cargo_a_definir:
            combo_cargo.set(cargo_a_definir)

        ttk.Label(row_frame, text="Horas Totais:").pack(side=tk.LEFT, padx=(0, 5))
        entry = ttk.Entry(row_frame, width=10)
        entry.pack(side=tk.LEFT, padx=5)
        entry.bind("<KeyRelease>", self._on_horas_keyrelease)
        if horas_a_definir is not None:
            entry.insert(0, f"{horas_a_definir:.2f}".replace(".", ","))

        action_buttons_frame = ttk.Frame(row_frame)
        action_buttons_frame.pack(side=tk.RIGHT, padx=(10, 0))

        aloc_widget_dict = {
            "type": "equipe",
            "combo_nome": combo_nome,
            "combo_cargo": combo_cargo,
            "entry": entry,
            "frame": row_frame,
            "disciplina": disciplina,
        }

        # New: "Change" Button
        change_btn = ttk.Button(
            action_buttons_frame,
            text="...",  # Or a pencil icon, etc.
            width=3,
            command=partial(
                self._enable_combobox_for_change,
                lote_nome,
                disciplina,
                aloc_widget_dict,
            ),
        )
        change_btn.pack(side=tk.LEFT, padx=(0, 2))
        aloc_widget_dict["change_button"] = change_btn  # Store reference

        remove_cmd = partial(
            self._remover_linha_alocacao, lote_nome, disciplina, aloc_widget_dict
        )
        remove_btn = ttk.Button(
            action_buttons_frame, text="X", width=3, command=remove_cmd
        )
        remove_btn.pack(side=tk.LEFT, padx=(0, 2))

        dividir_cmd = partial(
            self.dividir_tarefa, lote_nome, disciplina, aloc_widget_dict
        )
        dividir_btn = ttk.Button(
            action_buttons_frame, text="Dividir", width=7, command=dividir_cmd
        )
        dividir_btn.pack(side=tk.LEFT, padx=(5, 0))

        self.lote_widgets[lote_nome]["disciplinas"][disciplina][
            "alocacoes_widgets"
        ].append(aloc_widget_dict)

        # Bind <<ComboboxSelected>> to handle when a new selection is made
        combo_nome.bind(
            "<<ComboboxSelected>>",
            partial(self._on_combobox_selected, aloc_widget=aloc_widget_dict),
        )

        # Use postcommand to update options ONLY when the dropdown is about to appear
        combo_nome.config(
            postcommand=partial(self._populate_combobox_options, combo_nome, disciplina)
        )

        # Initial setup of the Combobox state and values
        # If a name is already set (from loading saved data), lock it immediately.
        # Otherwise, populate it with initial options.
        if combo_nome.get():
            # If there's a pre-set value, lock it.
            combo_nome.config(state="disabled")
            change_btn.config(state="normal")
            combo_nome.config(
                values=[combo_nome.get(), ""]
            )  # Only current selection + empty
        else:
            # If no value is set, it's ready for a new selection.
            combo_nome.config(state="readonly")
            change_btn.config(
                state="disabled"
            )  # Disable change button if nothing selected yet
            self._populate_combobox_options(
                combo_nome, disciplina
            )  # Populate initial options

    def _on_combobox_selected(self, event, aloc_widget):
        """
        Handler for when an item is selected from the employee combobox dropdown.
        It locks the combobox if a valid selection is made.
        """
        combo_nome = aloc_widget["combo_nome"]
        change_btn = aloc_widget["change_button"]
        selected_value = combo_nome.get()

        if selected_value:  # If a value was actually selected (not empty)
            # Re-lock the combobox
            combo_nome.config(state="disabled")
            change_btn.config(state="normal")
            # Set values to only include the selected one and empty to reinforce lock
            combo_nome.config(values=[selected_value, ""])
        else:  # If empty selection (clearing)
            combo_nome.config(state="readonly")  # Keep it open for new selection
            change_btn.config(state="disabled")  # No selected employee to change
            self._populate_combobox_options(
                combo_nome, aloc_widget["disciplina"]
            )  # Repopulate all available options

        self.processar_calculo(
            redirect_to_dashboard=False
        )  # Recalculate whenever selection changes

    def _criar_botao_adicionar_tarefa(self, parent_frame, lote_nome):
        """Cria e posiciona o botão '+ Adicionar Tarefa' no final do frame pai."""
        add_button_frame = ttk.Frame(parent_frame)
        add_button_frame.pack(fill=tk.X, padx=10, pady=20)

        # O comando agora chama a nova função de diálogo genérica
        cmd = partial(self._abrir_dialogo_adicionar_tarefa, lote_nome, parent_frame)

        # O texto do botão também foi generalizado
        ttk.Button(
            add_button_frame,
            text="+ Adicionar Tarefa",
            command=cmd,
            style="Accent.TButton",
        ).pack()

    def _calcular_horas_periodo(self, start_str, end_str):
        """Calcula o total de horas para um período e retorna como float, ou None se falhar."""
        try:
            horas_mes_str = self.horas_mes_entry.get()

            if not all([start_str, end_str, horas_mes_str]):
                return None

            inicio = datetime.datetime.strptime(start_str, "%d/%m/%Y")
            fim = datetime.datetime.strptime(end_str, "%d/%m/%Y")
            horas_mes = int(horas_mes_str)

            if fim < inicio:
                return None

            num_meses = (fim.year - inicio.year) * 12 + (fim.month - inicio.month) + 1
            total_horas_periodo = num_meses * horas_mes
            return float(total_horas_periodo)

        except (ValueError, TypeError):
            return None

    def _on_adicionar_alocacao_click(self, lote_nome, disc_nome):
        """
        Handler para o clique do botão '+ Alocar Equipe'.
        Pega as datas, calcula as horas e chama a função para criar a nova linha
        já com o valor das horas preenchido.
        """
        # Acessa os widgets da disciplina correta para pegar as datas
        disc_widgets = self.lote_widgets[lote_nome]["disciplinas"][disc_nome]
        start_str = disc_widgets["start_date"].get()
        end_str = disc_widgets["end_date"].get()

        # Usa a nova função para calcular as horas
        horas_calculadas = self._calcular_horas_periodo(start_str, end_str)

        # Chama a função que realmente cria os widgets na tela,
        # passando o valor calculado para o parâmetro 'horas_a_definir'.
        self.adicionar_alocacao_equipe(
            lote_nome, disc_nome, horas_a_definir=horas_calculadas
        )

    def _atualizar_horas_alocadas_por_disciplina(
        self, lote_nome, tarefa_nome, event=None
    ):
        if self._is_loading:
            return

        try:
            disc_widgets = self.lote_widgets[lote_nome]["disciplinas"][tarefa_nome]

            start_str = disc_widgets["start_date"].get()
            end_str = disc_widgets["end_date"].get()
            horas_mes_str = self.horas_mes_entry.get()

            if not all([start_str, end_str, horas_mes_str]):
                return  # Sai silenciosamente se as datas não estiverem prontas

            inicio = datetime.datetime.strptime(start_str, "%d/%m/%Y")
            fim = datetime.datetime.strptime(end_str, "%d/%m/%Y")
            horas_mes = int(horas_mes_str)

            if fim < inicio:
                return

            num_meses = (fim.year - inicio.year) * 12 + (fim.month - inicio.month) + 1
            total_horas_periodo = num_meses * horas_mes
            horas_formatadas = f"{total_horas_periodo:.2f}".replace(".", ",")

            # Itera sobre todas as alocações da tarefa
            for aloc in disc_widgets["alocacoes_widgets"]:
                if aloc["frame"].winfo_exists():
                    entry = aloc["entry"]
                    valor_atual_str = entry.get().strip()

                    # --- INÍCIO DA CORREÇÃO LÓGICA ---
                    deve_atualizar = False
                    if not valor_atual_str:
                        # Condição 1: O campo está literalmente vazio.
                        deve_atualizar = True
                    else:
                        try:
                            # Condição 2: O campo contém um número que é igual a zero.
                            valor_numerico = float(valor_atual_str.replace(",", "."))
                            if valor_numerico == 0:
                                deve_atualizar = True
                        except ValueError:
                            # Se o campo tiver um texto não numérico, não faz nada.
                            pass

                    if deve_atualizar:
                        entry.delete(0, tk.END)
                        entry.insert(0, horas_formatadas)
                    # --- FIM DA CORREÇÃO LÓGICA ---

        except (ValueError, TypeError, KeyError):
            # Captura qualquer erro de conversão de data, tipo ou busca de chave e ignora.
            pass

            # Em gui_app.py, na classe GuiApp

    def _recalcular_todas_as_horas_da_tarefa(self, lote_nome, tarefa_nome, event=None):
        """
        Calcula as horas do período e SOBRESCREVE o valor em todas as alocações da tarefa.
        Esta função é acionada pela mudança das datas.
        """
        if self._is_loading:
            return

        try:
            disc_widgets = self.lote_widgets[lote_nome]["disciplinas"][tarefa_nome]
            start_str = disc_widgets["start_date"].get()
            end_str = disc_widgets["end_date"].get()

            # Usa a função auxiliar que já temos para pegar o valor numérico das horas
            total_horas = self._calcular_horas_periodo(start_str, end_str)

            if total_horas is None:
                # Se o cálculo falhar (datas inválidas, etc.), não faz nada
                return

            horas_formatadas = f"{total_horas:.2f}".replace(".", ",")

            # Loop que sobrescreve TUDO
            for aloc in disc_widgets["alocacoes_widgets"]:
                if aloc["frame"].winfo_exists():
                    entry = aloc["entry"]
                    entry.delete(0, tk.END)
                    entry.insert(0, horas_formatadas)

        except (KeyError, ValueError):
            # Ignora erros se os widgets não forem encontrados ou as datas forem inválidas
            pass

    def _coletar_dados_da_ui(self):
        """
        Coleta TODOS os dados inseridos pelo usuário na interface e os retorna em um dicionário estruturado.
        Esta função lê de `self.lote_widgets` para lotes não-visíveis, e LÊ DIRETAMENTE DA UI
        para o lote atualmente visível.
        """
        try:
            dados = {"lotes": [], "config": {}, "funcionarios": []}
            dados["config"]["horas_mes"] = int(self.horas_mes_entry.get())
            dados["config"]["num_lotes"] = int(self.num_lotes_entry.get())
            dados["funcionarios"] = self.app_controller.funcionarios

            # Create a temporary deep copy of lote_widgets to work with,
            # so we don't modify the live UI references permanently.
            # Using a simple copy here, then handling the current lot's data explicitly.
            current_lote_data_snapshot = None
            if self.current_lote_name and self.current_lote_name in self.lote_widgets:
                # Capture the current lot's data from UI before processing all lots
                # This is similar to what _salvar_dados_lote_ui_para_memoria does,
                # but we'll use this data for the current cycle without affecting saved state.

                # Retrieve the currently active widgets for the visible lot
                active_lote_widgets_refs = self.lote_widgets[self.current_lote_name]

                # Temporarily store the data from current UI for processing
                temp_disciplines_data = {}
                if (
                    "scrollable_frame" in active_lote_widgets_refs
                    and active_lote_widgets_refs["scrollable_frame"].winfo_exists()
                ):
                    for child_frame in active_lote_widgets_refs[
                        "scrollable_frame"
                    ].winfo_children():
                        if isinstance(child_frame, ttk.LabelFrame):
                            nome_tarefa = child_frame.cget("text")

                            start_date_entry = None
                            end_date_entry = None
                            for widget in child_frame.winfo_children():
                                if (
                                    isinstance(widget, ttk.Frame)
                                    and widget.winfo_children()
                                ):
                                    if any(
                                        "Início" in c.cget("text")
                                        for c in widget.winfo_children()
                                        if hasattr(c, "cget")
                                    ):
                                        start_date_entry = widget.winfo_children()[1]
                                        end_date_entry = widget.winfo_children()[3]
                                        break

                            if not start_date_entry or not end_date_entry:
                                continue

                            disc_data_for_processing = {
                                "cronograma": {
                                    "inicio": start_date_entry.get(),
                                    "fim": end_date_entry.get(),
                                },
                                "alocacoes": [],
                            }

                            # Collect employee allocations
                            disc_info_for_active_widgets = active_lote_widgets_refs[
                                "disciplinas"
                            ].get(nome_tarefa)
                            if (
                                disc_info_for_active_widgets
                                and "alocacoes_widgets" in disc_info_for_active_widgets
                            ):
                                for aloc_widget_dict in disc_info_for_active_widgets[
                                    "alocacoes_widgets"
                                ]:
                                    if aloc_widget_dict["frame"].winfo_exists():
                                        nome = aloc_widget_dict["combo_nome"].get()
                                        cargo = aloc_widget_dict["combo_cargo"].get()
                                        horas_str = (
                                            aloc_widget_dict["entry"]
                                            .get()
                                            .replace(",", ".")
                                        )
                                        if nome and cargo and horas_str:
                                            try:
                                                horas = float(horas_str)
                                                disc_data_for_processing[
                                                    "alocacoes"
                                                ].append(
                                                    {
                                                        "funcionario": [nome, cargo],
                                                        "horas_totais": horas,
                                                    }
                                                )
                                            except ValueError:
                                                pass
                            temp_disciplines_data[nome_tarefa] = (
                                disc_data_for_processing
                            )

                # Collect subcontract allocations for current lot
                temp_subcontracts_data = []
                if "subcontratos_widgets" in active_lote_widgets_refs:
                    for sub_widget_dict in active_lote_widgets_refs[
                        "subcontratos_widgets"
                    ]:
                        if sub_widget_dict["frame"].winfo_exists():
                            sub_name = sub_widget_dict["name_entry"].get()
                            sub_start_date = sub_widget_dict["start_date_entry"].get()
                            sub_end_date = sub_widget_dict["end_date_entry"].get()
                            sub_hours_str = (
                                sub_widget_dict["hours_entry"].get().replace(",", ".")
                            )
                            if (
                                sub_name
                                and sub_start_date
                                and sub_end_date
                                and sub_hours_str
                            ):
                                try:
                                    sub_hours = float(sub_hours_str)
                                    temp_subcontracts_data.append(
                                        {
                                            "nome_subcontrato": sub_name,
                                            "cronograma": {
                                                "inicio": sub_start_date,
                                                "fim": sub_end_date,
                                            },
                                            "horas_totais": sub_hours,
                                        }
                                    )
                                except ValueError:
                                    pass

                current_lote_data_snapshot = {
                    "nome": self.current_lote_name,
                    "disciplinas": temp_disciplines_data,
                    "subcontratos": temp_subcontracts_data,
                    # Other lot properties might be needed if they are critical to calculation
                }

            # Now, iterate through all lots in self.lote_widgets to build the 'dados' dictionary
            for lote_name, lote_data_in_memory in self.lote_widgets.items():
                if lote_name == self.current_lote_name and current_lote_data_snapshot:
                    # Use the live snapshot data for the current lot
                    dados["lotes"].append(current_lote_data_snapshot)
                else:
                    # For other (non-current) lots, use the data already stored in memory
                    lote_dict = {"nome": lote_name, "disciplinas": {}}

                    for disc_name, disc_info_in_memory in lote_data_in_memory.get(
                        "disciplinas", {}
                    ).items():
                        disc_dict = {"cronograma": {}, "alocacoes": []}
                        disc_dict["cronograma"]["inicio"] = disc_info_in_memory.get(
                            "cronograma", {}
                        ).get("inicio", "")
                        disc_dict["cronograma"]["fim"] = disc_info_in_memory.get(
                            "cronograma", {}
                        ).get("fim", "")
                        disc_dict["alocacoes"] = disc_info_in_memory.get(
                            "alocacoes", []
                        )
                        lote_dict["disciplinas"][disc_name] = disc_dict

                    lote_dict["subcontratos"] = lote_data_in_memory.get(
                        "subcontratos", []
                    )
                    dados["lotes"].append(lote_dict)

            # Debug print
            import json

            return dados
        except (
            ValueError,
            KeyError,
            Exception,
        ) as e:  # Catch generic Exception for broader debugging
            import traceback

            traceback.print_exc()  # Print full traceback
            messagebox.showerror(
                "Erro de Entrada",
                f"Por favor, verifique os dados inseridos. Erro: {e}",
                parent=self,
            )
            return None

    # MODIFICADO: Coleta dados considerando a nova estrutura de funcionário.
    def salvar_estado(self):
        # NOVO: Salva o estado atual da UI em um arquivo JSON
        estado_para_salvar = self._coletar_dados_da_ui()
        if estado_para_salvar:
            try:
                with open(ESTADO_APP_FILE, "w", encoding="utf-8") as f:
                    json.dump(estado_para_salvar, f, ensure_ascii=False, indent=4)
                print("Estado salvo com sucesso.")
            except Exception as e:
                print(f"Erro ao salvar o estado: {e}")

    # MODIFICADO: Carrega o estado e recria as alocações corretamente.

    # Em gui_app.py, substitua esta função
    def carregar_ou_iniciar_ui(self):
        """
        Attempts to load the state from the JSON file. If successful, it populates
        the internal data structures and triggers rendering of the first lot.
        """
        self._is_loading = True  # Set the loading flag
        try:
            if not os.path.exists(ESTADO_APP_FILE):
                print("No saved state found. Starting with default UI.")
                self.gerar_abas_lotes()  # This will also set current_lote_name and trigger _on_lote_selected
                return

            with open(ESTADO_APP_FILE, "r", encoding="utf-8") as f:
                estado = json.load(f)

            # Load employees into the controller
            for func_data in estado.get("funcionarios", []):
                self.app_controller.adicionar_funcionario(*func_data)

            # Populate lote_widgets with loaded data. Don't create UI yet.
            self.lote_widgets.clear()
            lote_names_to_populate_combo = []
            for lote_data in estado.get("lotes", []):
                lote_name = lote_data["nome"]
                self.lote_widgets[lote_name] = {
                    "nome": lote_name,
                    "disciplinas": lote_data[
                        "disciplinas"
                    ],  # Store the loaded discipline structure (with data)
                    "data_entry_tab_frame": None,  # Will be created on demand by _renderizar_lote_na_ui
                    "dashboard_tab_frame": None,  # Will be created on demand
                    "scrollable_frame": None,  # Will be created on demand
                    "dash_tree": None,  # Will be created on demand
                }
                lote_names_to_populate_combo.append(lote_name)

            # Update the lot selection combobox
            self.lote_selection_combo.config(
                values=sorted(
                    lote_names_to_populate_combo, key=self._numerical_sort_key
                )
            )
            if lote_names_to_populate_combo:
                first_lote = sorted(
                    lote_names_to_populate_combo, key=self._numerical_sort_key
                )[0]
                self.lote_selection_combo.set(first_lote)
                self.current_lote_name = first_lote  # Set the initial visible lot

                # Trigger the _on_lote_selected to render the first lot's UI and process it
                self._on_lote_selected(None)  # Pass None as event object

            # Render general configurations
            config = estado.get("config", {})
            self.horas_mes_entry.delete(0, tk.END)
            self.horas_mes_entry.insert(0, str(config.get("horas_mes", 160)))
            self.num_lotes_entry.delete(0, tk.END)
            self.num_lotes_entry.insert(0, str(config.get("num_lotes", 1)))

            print("State loaded successfully.")
            # process_calculo is already called by _on_lote_selected

        except Exception as e:
            # Error block (unchanged)
            import traceback

            messagebox.showerror(
                "Erro ao Carregar",
                f"Não foi possível carregar o estado salvo: {e}\n\nIniciando com uma nova sessão.",
                parent=self,
            )
            traceback.print_exc()
            self.lote_widgets.clear()
            # No longer need to forget notebook tabs here, as they are now managed per-lot dynamically
            self.gerar_abas_lotes()  # Revert to default UI in case of error
        finally:
            self._is_loading = False  # Reset the loading flag

    def _renderizar_estado_na_ui(self, estado_completo):
        """
        Função MESTRE de renderização. Pega um dicionário de 'estado' e
        desenha/atualiza toda a interface para corresponder a ele.
        Esta função é chamada uma única vez no carregamento inicial (agora por carregar_ou_iniciar_ui).
        """
        self._is_loading = True
        try:
            # 1. Limpa o combobox de lotes e o frame de conteúdo
            self.lote_selection_combo.set("")
            self.lote_selection_combo.config(values=[])

            # Limpa o notebook interno do lote
            for tab in self.lote_inner_notebook.tabs():
                self.lote_inner_notebook.forget(tab)
            # Os frames current_lote_content_frame e current_lote_dashboard_frame
            # não são destruídos, apenas seu conteúdo é limpo.
            for widget in self.current_lote_content_frame.winfo_children():
                widget.destroy()
            self.current_lote_name = None

            # 2. Renderiza as configurações gerais
            config = estado_completo.get("config", {})
            self.horas_mes_entry.delete(0, tk.END)
            self.horas_mes_entry.insert(0, str(config.get("horas_mes", 160)))
            self.num_lotes_entry.delete(0, tk.END)
            self.num_lotes_entry.insert(0, str(config.get("num_lotes", 1)))

            # 3. Preenche o dicionário `self.lote_widgets` e o combobox de lotes
            self.lote_widgets.clear()
            lote_names = []
            for lote_data in estado_completo.get("lotes", []):
                lote_name = lote_data["nome"]
                self.lote_widgets[lote_name] = {
                    "nome": lote_name,
                    "disciplinas": lote_data[
                        "disciplinas"
                    ],  # Armazena a estrutura de disciplinas carregada (com dados)
                    "dash_tree": self._criar_treeview(
                        self.current_lote_dashboard_frame
                    ),  # Usa o frame do dashboard do lote
                    "scrollable_frame": None,  # Será preenchido na renderização específica do lote
                }
                lote_names.append(lote_name)

            self.lote_selection_combo.config(values=sorted(lote_names))

            # Seleciona o primeiro lote para exibição
            if lote_names:
                first_lote_name = sorted(lote_names)[0]
                self.lote_selection_combo.set(first_lote_name)
                self.current_lote_name = first_lote_name
                self._renderizar_lote_na_ui(first_lote_name)

        finally:
            self.update_idletasks()
            self._is_loading = False

    def processar_calculo(self, redirect_to_dashboard=True):
        """Process all collected data and update the UI with the results,
        including consolidated and lot-specific dashboards.
        """
        dados_coletados = self._coletar_dados_da_ui()
        if not dados_coletados:
            return
        try:
            horas_mes = dados_coletados["config"]["horas_mes"]

            # This calls the App controller to perform the main calculations
            self.app_controller.processar_portfolio(dados_coletados["lotes"], horas_mes)

            # Update the consolidated dashboard (this remains correct)
            self._preencher_treeview(
                self.dash_tree, self.app_controller.get_ultimo_df_consolidado()
            )

            # Update the dashboard for the CURRENTLY ACTIVE lot
            # In this reverted setup, self.current_lote_dash_tree is the correct object for the active lot.
            if self.current_lote_name:
                df_lote_ativo = self.app_controller.get_ultimo_dashboards_lotes().get(
                    self.current_lote_name, pd.DataFrame()
                )
                # THIS IS THE CRUCIAL LINE RE-ADDED:
                self._preencher_treeview(self.current_lote_dash_tree, df_lote_ativo)

            # Handle the composition alert (this part remains unchanged)
            porcentagem_global, detalhes_lotes = (
                self.app_controller.portfolio.verificar_porcentagem_horas_cargo(
                    cargo_alvo=self.app_controller.get_cargo_estagiario()
                )
            )
            alerta_composicao = {"global": porcentagem_global, "lotes": detalhes_lotes}

            if alerta_composicao and (
                alerta_composicao["global"] >= 50.0
                or any(p >= 50.0 for p in alerta_composicao["lotes"].values())
            ):
                self.mostrar_alerta_composicao(alerta_composicao)

            if redirect_to_dashboard:
                self.notebook.select(0)  # Selects the consolidated dashboard tab

        except Exception as e:
            messagebox.showerror(
                "Erro Inesperado", f"Ocorreu um erro no processamento: {e}", parent=self
            )

    # Helper function (if it doesn't exist, you'll need to add it to GuiApp)
    # This function centralizes the logic for showing the composition alert
    def mostrar_alerta_composicao(self, alerta_composicao):
        """Displays a messagebox alert about team composition."""
        porcentagem_global = alerta_composicao.get("global", 0.0)
        detalhes_lotes = alerta_composicao.get("lotes", {})

        lotes_excedidos = [
            lote for lote, perc in detalhes_lotes.items() if perc >= 50.0
        ]

        if porcentagem_global >= 50.0 or lotes_excedidos:
            if porcentagem_global >= 50.0:
                mensagem_alerta = (
                    f"Atenção: A participação global de Estagiários/Projetistas é de "
                    f"{porcentagem_global:.1f}%, excedendo o limite de 50%."
                ).replace(".", ",")
            else:
                mensagem_alerta = (
                    f"Atenção: Um ou mais lotes excederam o limite individual de 50% de horas "
                    f"para Estagiários/Projetistas (Global: {porcentagem_global:.1f}%)."
                ).replace(".", ",")

            mensagem_alerta += "\n\nDetalhes da Participação por Lote:"

            lotes_ordenados = sorted(
                detalhes_lotes.items(), key=lambda item: item[1], reverse=True
            )

            for nome_lote, percent_lote in lotes_ordenados:
                marcador = "  <<-- LIMITE EXCEDIDO!" if percent_lote >= 50.0 else ""
                mensagem_alerta += (
                    f"\n- Lote {nome_lote}: {percent_lote:.1f}%{marcador}".replace(
                        ".", ","
                    )
                )

            messagebox.showwarning(
                "Alerta de Composição da Equipe", mensagem_alerta, parent=self
            )

    # ... O restante do arquivo (dividir_tarefa, exportar_para_excel, etc.) pode permanecer o mesmo por enquanto
    # ... ou ser ajustado se a função dividir tarefa precisar de lógica de categoria
    def dividir_tarefa(self, lote_nome, disciplina, aloc_origem_widget):
        """
        Função com lógica corrigida para dividir uma tarefa.
        Ela divide as horas e cria uma nova linha de alocação para o usuário preencher.
        """
        try:
            horas_str = aloc_origem_widget["entry"].get().replace(",", ".")
            horas_originais = float(horas_str or 0.0)

            if horas_originais <= 0:
                messagebox.showwarning(
                    "Aviso",
                    "A tarefa precisa ter horas positivas para ser dividida.",
                    parent=self,
                )
                return

            horas_divididas = horas_originais / 2.0

            # Atualiza a entrada original com metade das horas
            aloc_origem_widget["entry"].delete(0, tk.END)
            aloc_origem_widget["entry"].insert(
                0, f"{horas_divididas:.2f}".replace(".", ",")
            )

            # Adiciona uma nova linha de alocação com a outra metade das horas
            # e o mesmo cargo da linha original. O usuário selecionará o funcionário.
            cargo_original = aloc_origem_widget["combo_cargo"].get()
            self.adicionar_alocacao_equipe(
                lote_nome,
                disciplina,
                cargo_a_definir=cargo_original,
                horas_a_definir=horas_divididas,
            )

            # Adiciona os dados da nova alocação à lista de dados persistidos
            nova_aloc_data_para_memoria = {
                "funcionario": [
                    "",
                    cargo_original,
                ],  # O funcionário será definido pelo usuário
                "horas_totais": horas_divididas,
            }
            self.lote_widgets[lote_nome]["disciplinas"][disciplina]["alocacoes"].append(
                nova_aloc_data_para_memoria
            )

        except (ValueError, KeyError) as e:
            messagebox.showerror(
                "Erro ao Dividir",
                f"Não foi possível dividir a tarefa: {e}",
                parent=self,
            )

    def _numerical_sort_key(self, item):
        try:
            return int(item)
        except ValueError:
            return item

    def gerar_abas_lotes(self, popular_defaults=True):
        """
        This function now updates the list of lots in the combobox and
        recreates the UI for the initially selected lot.
        """
        # Save the state of the current visible lot before generating new lots
        if self.current_lote_name:
            self._salvar_dados_lote_ui_para_memoria(self.current_lote_name)

        try:
            num_lotes = int(self.num_lotes_entry.get())
            if num_lotes <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Erro", "Insira um número válido e positivo de lotes.")
            return

        # Clear the lot data in memory and the selection combobox
        self.lote_widgets.clear()
        self.lote_selection_combo.set("")
        self.lote_selection_combo.config(values=[])

        # Clear the inner notebook tabs and destroy existing widgets within the content frames
        for tab in self.lote_inner_notebook.tabs():
            self.lote_inner_notebook.forget(tab)
        for widget in self.current_lote_content_frame.winfo_children():
            widget.destroy()
        self.current_lote_dash_tree.delete(*self.current_lote_dash_tree.get_children())

        self.current_lote_name = None

        lote_names = []
        for i in range(num_lotes):
            lote_nome = str(i + 1)
            lote_names.append(lote_nome)

            # Initialize the data structure for the new lot (data only)
            self.lote_widgets[lote_nome] = {
                "nome": lote_nome,
                "disciplinas": {},  # Will be filled below or on load
            }

            if popular_defaults:
                disciplinas_iniciais = self.app_controller.get_disciplinas()
                for disc in disciplinas_iniciais:
                    # Add discipline to the lot's data dictionary
                    self.lote_widgets[lote_nome]["disciplinas"][disc] = {
                        "cronograma": {
                            "inicio": "",
                            "fim": "",
                        },
                        "alocacoes": [],
                    }
                    # Populate with default allocations (data, not UI widgets)
                    cargos_disponiveis = self.app_controller.get_cargos_disponiveis()
                    for cargo_padrao in cargos_disponiveis:
                        self.lote_widgets[lote_nome]["disciplinas"][disc][
                            "alocacoes"
                        ].append(
                            {
                                "funcionario": [
                                    "",
                                    cargo_padrao,
                                ],
                                "horas_totais": 0.0,
                            }
                        )

        # Update the combobox with lot names
        self.lote_selection_combo.config(
            values=sorted(lote_names, key=self._numerical_sort_key)
        )
        if lote_names:
            first_lote = sorted(lote_names)[0]
            self.lote_selection_combo.set(first_lote)
            self.current_lote_name = first_lote
            # Render the UI for the first lot after generating
            self._renderizar_lote_na_ui(first_lote)

        self.atualizar_lista_funcionarios()

    def _popular_lotes_com_defaults(self):
        """
        Esta função não é mais diretamente chamada para popular a UI,
        mas sim para preencher a estrutura de dados interna do lote.
        A renderização da UI ocorre via `_renderizar_lote_na_ui`.
        """
        # A lógica de popular com defaults foi movida para `gerar_abas_lotes`
        # para inicializar a estrutura de dados interna do lote.
        pass

    def criar_conteudo_aba_entrada(self, parent_tab, lote_name, popular_defaults=True):
        # Esta parte inicial que cria a área de rolagem está correta.
        canvas = tk.Canvas(parent_tab, highlightthickness=0, bg="#FFFFFF")
        scrollbar = ttk.Scrollbar(parent_tab, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, style="Modern.TFrame")

        scrollable_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        def _on_mousewheel(event):
            canvas.yview_scroll(-1 * (event.delta // 120), "units")

        def _bind_scroll(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_scroll(event):
            canvas.unbind_all("<MouseWheel>")

        canvas.bind("<Enter>", _bind_scroll)
        canvas.bind("<Leave>", _unbind_scroll)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Se for um lote novo, cria as disciplinas padrão e o botão no final
        # A lógica de criação das disciplinas e alocações padrão está agora em gerar_abas_lotes
        # e é aplicada à estrutura de dados do lote. _renderizar_lote_na_ui então constrói a UI a partir dela.

        # O botão "+ Adicionar Tarefa" é adicionado por _renderizar_lote_na_ui
        # após popular as tarefas existentes.

        # Adicionamos um retorno para que a função de carregamento possa obter a referência ao frame
        return scrollable_frame

    def _populate_combobox_options(self, combobox_widget, nome_da_tarefa):
        """
        Populates the values of a given employee combobox with available, non-used employees.
        This function does NOT handle locking/unlocking the combobox state.
        It expects the combobox to be in an editable state (e.g., 'readonly').
        """
        if not combobox_widget.winfo_exists():
            return

        selecao_atual = (
            combobox_widget.get()
        )  # Keep current selection attempt to restore

        nomes_usados_no_lote = set()
        lote_data = self.lote_widgets.get(self.current_lote_name, {})

        if "disciplinas" in lote_data:
            for disc_data_value in lote_data["disciplinas"].values():
                for aloc_widget in disc_data_value.get("alocacoes_widgets", []):
                    combo_iter = aloc_widget.get("combo_nome")

                    if (
                        combo_iter
                        and combo_iter.winfo_exists()
                        and combo_iter is not combobox_widget
                    ):
                        selecao = combo_iter.get()
                        if selecao:
                            nomes_usados_no_lote.add(selecao)

        lista_nomes = self.app_controller.get_funcionarios_para_tarefa(nome_da_tarefa)

        opcoes_disponiveis = [
            nome for nome in lista_nomes if nome not in nomes_usados_no_lote
        ]

        # Always include the current selection if it's a valid employee (for smooth updates)
        if (
            selecao_atual
            and selecao_atual in self.app_controller.get_todos_os_funcionarios()
            and selecao_atual not in opcoes_disponiveis
        ):
            opcoes_disponiveis.insert(0, selecao_atual)

        # Set the values for the combobox.
        combobox_widget.config(values=[""] + sorted(opcoes_disponiveis))

        # Restore the selection if it's still valid
        if selecao_atual in combobox_widget["values"]:
            combobox_widget.set(selecao_atual)
        # Else, if it's no longer valid (e.g., employee removed), the field will stay empty
        # or the previous (now invalid) value might remain until a new selection.
        # This is where the _on_combobox_selected logic helps.

    def _criar_frame_tarefa(
        self, parent_frame, lote_name, nome_tarefa, popular_defaults=True
    ):
        """
        Cria um frame individual para uma tarefa (disciplina/subcontrato) dentro de um lote.
        Este método é chamado para RECONSTRUIR a UI do lote selecionado.
        """
        disc_frame = ttk.LabelFrame(parent_frame, text=nome_tarefa, padding=10)
        disc_frame.pack(fill=tk.X, expand=True, padx=10, pady=(10, 5))

        disc_frame.columnconfigure(0, weight=1)
        date_frame = ttk.Frame(disc_frame)
        date_frame.grid(row=0, column=0, sticky="w", pady=(0, 5))

        ttk.Label(date_frame, text="Início (DD/MM/AAAA):").pack(side=tk.LEFT)
        start_date_entry = ttk.Entry(date_frame, width=15)
        start_date_entry.pack(side=tk.LEFT, padx=5)
        start_date_entry.bind("<KeyRelease>", self._on_data_keyrelease)

        ttk.Label(date_frame, text="Fim (DD/MM/AAAA):").pack(side=tk.LEFT, padx=10)
        end_date_entry = ttk.Entry(date_frame, width=15)
        end_date_entry.pack(side=tk.LEFT, padx=5)
        end_date_entry.bind("<KeyRelease>", self._on_data_keyrelease)

        aloc_frame = ttk.Frame(disc_frame)
        aloc_frame.grid(row=1, column=0, sticky="ew", pady=5)

        # As referências dos widgets ativos para este lote/disciplina/tarefa
        # são armazenadas aqui para que _salvar_dados_lote_ui_para_memoria possa acessá-las.
        # As alocações_widgets serão preenchidas por adicionar_alocacao_equipe
        self.lote_widgets[lote_name]["disciplinas"][nome_tarefa][
            "start_date"
        ] = start_date_entry
        self.lote_widgets[lote_name]["disciplinas"][nome_tarefa][
            "end_date"
        ] = end_date_entry
        self.lote_widgets[lote_name]["disciplinas"][nome_tarefa][
            "alocacoes_frame"
        ] = aloc_frame
        self.lote_widgets[lote_name]["disciplinas"][nome_tarefa]["frame"] = disc_frame
        # Garante que alocacoes_widgets seja uma lista (mesmo que vazia inicialmente)
        if (
            "alocacoes_widgets"
            not in self.lote_widgets[lote_name]["disciplinas"][nome_tarefa]
        ):
            self.lote_widgets[lote_name]["disciplinas"][nome_tarefa][
                "alocacoes_widgets"
            ] = []

        calculo_cmd = partial(
            self._recalcular_todas_as_horas_da_tarefa, lote_name, nome_tarefa
        )
        start_date_entry.bind("<FocusOut>", calculo_cmd)
        end_date_entry.bind("<FocusOut>", calculo_cmd)

        # Se popular_defaults é True, isso significa que estamos criando um NOVO lote/tarefa do zero
        # (ex: ao clicar em "Gerar Lotes" com num_lotes > lotes existentes, ou "+ Adicionar Tarefa")
        # Neste caso, as alocações padrão para esta tarefa (disciplina) são adicionadas.
        if popular_defaults:
            # Se a tarefa já tem alocações nos dados (ex: carregadas do JSON), não recria
            # Se for uma tarefa "Nova" ou "Outros..." adicionada, a lista de alocações estará vazia
            # e poderá ser populada aqui com defaults ou deixada vazia.
            # A lógica de popular com defaults para novas tarefas deve estar em _abrir_dialogo_adicionar_tarefa.
            pass  # A adição das alocações é agora tratada pelo _renderizar_lote_na_ui ou _abrir_dialogo_adicionar_tarefa

        cmd_add_func = partial(
            self._on_adicionar_alocacao_click, lote_name, nome_tarefa
        )
        ttk.Button(disc_frame, text="+ Alocar Equipe", command=cmd_add_func).grid(
            row=2, column=0, sticky="w", pady=5
        )

    def _preencher_treeview(self, tree, df_relatorio):
        # Clear existing items
        tree.delete(*tree.get_children())

        # If DataFrame is empty, just clear and return
        if df_relatorio.empty:
            tree["columns"] = []  # Clear columns if no data
            return

        # Ensure "H.total" and "Status" are always the last columns if they exist
        cols = list(df_relatorio.columns)
        if "H.total" in cols:
            cols.remove("H.total")
            cols.append("H.total")
        if "Status" in cols:
            cols.remove("Status")
            cols.append("Status")

        # Set columns for the Treeview
        tree["columns"] = cols
        tree["show"] = "headings"

        # Configure tags (assuming these styles are defined)
        tree.tag_configure("excedido", background="#ffdddd", foreground="red")
        tree.tag_configure(
            "header", background="#e0e0e0", font=("Segoe UI", 10, "bold")
        )

        # Configure column headings and initial width
        font = tkFont.Font(family="Segoe UI", size=10)  # Define font once
        for col_name in cols:
            tree.heading(col_name, text=col_name, anchor=tk.CENTER)

            # Dynamic column width calculation - slightly adjusted for better spacing
            # Use max of column name length and max data length
            max_len = max(
                len(str(col_name)),  # Column header length
                df_relatorio[col_name]
                .astype(str)
                .apply(len)
                .max(),  # Max data length in column
            )
            width = (
                max_len * font.measure("M") // font.measure("MM") + 20
            )  # Estimate width, add padding
            tree.column(
                col_name, width=max(width, 70), anchor=tk.CENTER
            )  # Minimum width of 70 pixels

        # Populate the Treeview with data
        last_disc = None
        for _, row in df_relatorio.iterrows():
            # Check if 'Disciplina' column exists for grouping
            if "Disciplina" in row and row["Disciplina"] != last_disc:
                last_disc = row["Disciplina"]
                # Insert a header row for the discipline
                tree.insert(
                    "",
                    "end",
                    values=[last_disc.upper()] + [""] * (len(cols) - 1),
                    tags=("header",),
                )

            # Determine tags for the row (e.g., "excedido")
            tags = ()
            if "Status" in row and row["Status"] == "Alocação Excedida":
                tags = ("excedido",)

            # Format numeric values to two decimal places with comma as separator
            formatted_values = []
            for col in cols:
                value = row[col]
                # Check if the value is numeric (int or float) and not part of the initial columns like 'Disciplina', 'Cargo', 'Funcionário'
                if isinstance(value, (int, float)) and col not in [
                    "Disciplina",
                    "Cargo",
                    "Funcionário",
                    "Status",
                ]:
                    formatted_values.append(f"{value:.2f}".replace(".", ","))
                else:
                    formatted_values.append(value)

            # Insert the actual data row
            tree.insert("", "end", values=formatted_values, tags=tags)

        # Ensure visibility of vertical scrollbar if content exceeds view
        tree.update_idletasks()  # Ensure sizes are calculated

    def atualizar_dashboards(
        self,
        df_consolidado,
        dashboards_lotes,
        detalhes_tarefas=None,
        alerta_composicao=None,
    ):
        # This function no longer explicitly updates lot-specific dashboards,
        # as processar_calculo directly calls _preencher_treeview for them.
        # This function primarily handles the consolidated dashboard and alerts.
        self._preencher_treeview(self.dash_tree, df_consolidado)

        # The individual lot dashboard tree is updated directly in processar_calculo
        # so this part is removed here.
        # if self.current_lote_name:
        #     df_lote_ativo = dashboards_lotes.get(self.current_lote_name, pd.DataFrame())
        #     self._preencher_treeview(self.current_lote_dash_tree, df_lote_ativo)

        if alerta_composicao:
            porcentagem_global = alerta_composicao.get("global", 0.0)
            detalhes_lotes = alerta_composicao.get("lotes", {})

            lotes_excedidos = [
                lote for lote, perc in detalhes_lotes.items() if perc >= 50.0
            ]

            if porcentagem_global >= 50.0 or lotes_excedidos:
                if porcentagem_global >= 50.0:
                    mensagem_alerta = (
                        f"Atenção: A participação global de Estagiários/Projetistas é de "
                        f"{porcentagem_global:.1f}%, excedendo o limite de 50%."
                    ).replace(".", ",")
                else:
                    mensagem_alerta = (
                        f"Atenção: Um ou mais lotes excederam o limite individual de 50% de horas "
                        f"para Estagiários/Projetistas (Global: {porcentagem_global:.1f}%)."
                    ).replace(".", ",")

                mensagem_alerta += "\n\nDetalhes da Participação por Lote:"

                lotes_ordenados = sorted(
                    detalhes_lotes.items(), key=lambda item: item[1], reverse=True
                )

                for nome_lote, percent_lote in lotes_ordenados:
                    marcador = "  <<-- LIMITE EXCEDIDO!" if percent_lote >= 50.0 else ""
                    mensagem_alerta += (
                        f"\n- Lote {nome_lote}: {percent_lote:.1f}%{marcador}".replace(
                            ".", ","
                        )
                    )

                messagebox.showwarning(
                    "Alerta de Composição da Equipe", mensagem_alerta, parent=self
                )

    def exportar_para_excel(self):
        df_consolidado = self.app_controller.get_ultimo_df_consolidado()
        dashboards_lotes = self.app_controller.get_ultimo_dashboards_lotes()
        if df_consolidado is None or df_consolidado.empty:
            messagebox.showwarning(
                "Aviso",
                "Não há dados consolidados para exportar. Calcule a alocação primeiro.",
            )
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivo Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")],
            title="Salvar Relatório de Alocação",
            initialfile=f"Relatorio_Alocacao_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
        )
        if not filepath:
            return

        try:
            with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                # 1. Escreve a planilha consolidada e cria o gráfico para ela
                self._write_styled_df_to_excel(writer, "Consolidado", df_consolidado)
                self._create_chart_for_sheet(writer, "Consolidado", df_consolidado)

                # 2. Escreve as planilhas de lote e seus gráficos
                if dashboards_lotes:
                    for nome_lote, df_lote in dashboards_lotes.items():
                        if not df_lote.empty:
                            sheet_name = f"Lote {nome_lote}"
                            self._write_styled_df_to_excel(writer, sheet_name, df_lote)
                            self._create_chart_for_sheet(writer, sheet_name, df_lote)

                # 3. Remove a planilha padrão APENAS DEPOIS de criar as novas
                if "Sheet" in writer.book.sheetnames:
                    writer.book.remove(writer.book["Sheet"])

            messagebox.showinfo(
                "Sucesso", f"Relatório exportado com sucesso para:\n{filepath}"
            )
        except Exception as e:
            messagebox.showerror(
                "Erro na Exportação", f"Ocorreu um erro ao salvar o arquivo:\n{e}"
            )

    def _write_styled_df_to_excel(self, writer, sheet_name, df):
        # Remove a coluna de status que não é necessária no Excel
        df_para_exportar = df.drop(columns=["Status"], errors="ignore")
        if df_para_exportar.empty:
            return

        # Ordena por Disciplina para o agrupamento
        df_sorted = df_para_exportar.sort_values(
            by=["Disciplina", "Cargo", "Funcionário"]
        )

        # Escreve os dados na planilha
        df_sorted.to_excel(
            writer, sheet_name=sheet_name, index=False, header=False, startrow=1
        )
        ws = writer.sheets[sheet_name]
        ws.sheet_view.showGridLines = False

        # --- ESTILOS PADRONIZADOS ---
        header_font = Font(name="Arial", bold=True, color="000000", size=11)
        header_fill = PatternFill(
            start_color="E68A00", end_color="E68A00", fill_type="solid"
        )
        category_font = Font(name="Arial", bold=True, size=12)
        category_fill = PatternFill(
            start_color="E6AF60", end_color="E6AF60", fill_type="solid"
        )  # Azul claro
        data_font = Font(name="Arial", size=10)
        thin_side = Side(border_style="thin", color="BFBFBF")
        full_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )

        # Escreve e formata o cabeçalho
        headers = list(df_sorted.columns)
        for c_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill

        # Agrupamento por Disciplina
        last_discipline = None
        current_row_offset = 0
        # Itera sobre as linhas de dados para inserir os cabeçalhos de categoria
        for r_idx, row in enumerate(
            dataframe_to_rows(df_sorted, index=False, header=False), 2
        ):
            if row[0] != last_discipline:
                ws.insert_rows(r_idx + current_row_offset)

                cat_cell = ws.cell(
                    row=r_idx + current_row_offset, column=1, value=row[0]
                )
                cat_cell.font = category_font
                cat_cell.fill = category_fill
                ws.merge_cells(
                    start_row=r_idx + current_row_offset,
                    start_column=1,
                    end_row=r_idx + current_row_offset,
                    end_column=len(headers),
                )

                last_discipline = row[0]
                current_row_offset += 1

        # Aplica bordas e fontes a todas as células com conteúdo
        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row_cells:
                if cell.value is not None or cell.style == "input_style_resumo":
                    cell.border = full_border
                    if not cell.font.bold:
                        cell.font = data_font

        # Ajusta a largura das colunas
        for i, col_name in enumerate(headers, 1):
            column_letter = openpyxl.utils.get_column_letter(i)
            max_len = max(df_sorted[col_name].astype(str).map(len).max(), len(col_name))
            ws.column_dimensions[column_letter].width = max(max_len + 4, 12)

    def _create_chart_for_sheet(self, writer, sheet_name, df):
        from openpyxl.chart.axis import ChartLines
        from openpyxl.chart.text import RichText
        from openpyxl.drawing.text import (
            CharacterProperties,
            Paragraph,
            ParagraphProperties,
            RichTextProperties,
        )
        from openpyxl.drawing.line import LineProperties
        from openpyxl.drawing.fill import SolidColorFillProperties

        if df.empty:
            print(
                f"DEBUG: _create_chart_for_sheet: DataFrame for {sheet_name} is empty. Skipping chart creation."
            )
            return

        ws = writer.book[sheet_name]

        # Identify month columns in the DataFrame's current state (e.g., 'Jan/24')
        # These are usually strings formatted with commas
        month_cols_formatted = [
            col for col in df.columns if "/" in col and len(col) == 6  # e.g., 'Jan/24'
        ]

        if not month_cols_formatted:
            print(
                f"DEBUG: _create_chart_for_sheet: No formatted month columns found for {sheet_name}. Skipping chart."
            )
            return

        # Prepare data for charting: sum up the decimal values per month
        # We need to convert these columns to numeric values first.
        chart_data = {}
        for col in month_cols_formatted:
            try:
                # 1. Ensure the column is treated as string for .str accessor
                # 2. Replace comma with dot
                # 3. Convert to numeric. 'errors="coerce"' will turn unparseable values into NaN.
                numeric_series = pd.to_numeric(
                    df[col].astype(str).str.replace(",", "."), errors="coerce"
                )
                chart_data[col] = numeric_series.sum()
            except Exception as e:
                print(
                    f"ERROR: _create_chart_for_sheet: Failed to process column '{col}' for chart in sheet '{sheet_name}': {e}"
                )
                chart_data[col] = 0.0  # Default to 0 if there's an issue

        monthly_decimal_totals = pd.Series(chart_data)

        if monthly_decimal_totals.sum() == 0:
            print(
                f"DEBUG: _create_chart_for_sheet: Sum of monthly totals is zero for {sheet_name}. Skipping chart creation."
            )
            return

        # Sort months for the chart (important for chronological display)
        # Assuming format "Mon/YY" e.g., "Jan/24"
        try:
            sorted_months = sorted(
                monthly_decimal_totals.index,
                key=lambda x: datetime.datetime.strptime(x, "%b/%y"),
            )
            monthly_decimal_totals = monthly_decimal_totals[sorted_months]
        except ValueError as e:
            print(
                f"WARNING: _create_chart_for_sheet: Could not sort month columns chronologically for {sheet_name}: {e}"
            )
            # Fallback to default sorting if date parsing fails
            monthly_decimal_totals = monthly_decimal_totals.sort_index()

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = f"Histograma de pessoal - {sheet_name}\n(Homen/mês)"
        chart.title.layout = Layout(manualLayout=ManualLayout(y=0, yMode="edge"))
        chart.legend = None

        chart.y_axis.majorGridlines = None
        axis_line_props = LineProperties(solidFill="000000")
        chart.y_axis.spPr = GraphicalProperties(ln=axis_line_props)
        chart.x_axis.spPr = GraphicalProperties(ln=axis_line_props)
        chart.x_axis.delete = False
        chart.y_axis.delete = False

        chart.layout = Layout(
            manualLayout=ManualLayout(
                y=0.15,
                h=0.75,
            )
        )

        # Write chart data to a temporary hidden area in the sheet
        # Start from a column far to the right to avoid interfering with main data
        data_start_col = ws.max_column + 2  # Start 2 columns after the last used column
        chart_data_start_row = 1  # Start from row 1 for header

        # Write header for chart data
        ws.cell(row=chart_data_start_row, column=data_start_col, value="Mês")
        ws.cell(
            row=chart_data_start_row, column=data_start_col + 1, value="Total Decimal"
        )

        # Write the monthly totals
        for r_idx, (month, total) in enumerate(
            monthly_decimal_totals.items(), start=chart_data_start_row + 1
        ):
            ws.cell(row=r_idx, column=data_start_col, value=month)
            ws.cell(row=r_idx, column=data_start_col + 1, value=total)

        # Define chart references
        values = Reference(
            ws,
            min_col=data_start_col + 1,
            min_row=chart_data_start_row,  # Include header row for series title
            max_row=chart_data_start_row + len(monthly_decimal_totals),
        )
        cats = Reference(
            ws,
            min_col=data_start_col,
            min_row=chart_data_start_row + 1,  # Exclude header row for categories
            max_row=chart_data_start_row + len(monthly_decimal_totals),
        )

        series = Series(
            values, title_from_data=True
        )  # Use title_from_data=True if values include header
        series.graphicalProperties.solidFill = "E68A00"
        series.graphicalProperties.line.solidFill = "E68A00"

        chart.append(series)
        chart.set_categories(cats)

        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True
        chart.dLbls.showSerName = False
        chart.dLbls.showCatName = False
        chart.dLbls.showLegendKey = False
        chart.dLbls.position = "outEnd"
        chart.dLbls.numFmt = "0.00"

        # Position the chart below the main data table
        # Find the last row of your main data table + some buffer
        last_data_row = ws.max_row + 5  # Adjust buffer as needed
        chart_anchor_cell = f"A{last_data_row}"  # Anchor chart starting at column A

        ws.add_chart(chart, chart_anchor_cell)


# --- INÍCIO DAS MODIFICAÇÕES ---

# 2. Mover o título para cima ajustando o layout da área de plotagem
# Isso move a área do gráfico (as barras) para baixo, dando mais espaço para o título.
# chart.x_axis.txPr = RichText(
# bodyPr=RichTextProperties(
# rot=-450000, # Rotação de -45 graus (multiplicado por 100000)
# kern=1, # Ajuste de espaçamento (opcional, experimente)
# You might also need to set anchor or anchorCtr depending on exact positioning
# anchor="ctr", # Center anchor for the text
# anchorCtr=True, # Center anchor for the text
# ),
# p=[
# Paragraph(
# pPr=ParagraphProperties(defRPr=CharacterProperties()),
# endParaRPr=CharacterProperties(),
# )
# ],
# )
